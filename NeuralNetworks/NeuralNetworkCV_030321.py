#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# For handling data
import pandas as pd
import numpy as np
from datetime import datetime
from sklearn.preprocessing import OrdinalEncoder
from sklearn.utils import class_weight
from statistics import mean

# For feature selection
from sklearn.feature_selection import SelectKBest, chi2
from xgboost import XGBClassifier
from matplotlib import pyplot
from xgboost import plot_importance
from sklearn.metrics import accuracy_score
from sklearn.model_selection import train_test_split
from sklearn.feature_selection import mutual_info_classif
from sklearn import preprocessing
from sklearn.feature_selection import f_classif
# For balancing batches
from imblearn.keras import BalancedBatchGenerator
from imblearn.over_sampling import RandomOverSampler

# For NN and tuning
import tensorflow as tf
import kerastuner
from kerastuner.tuners import Hyperband, BayesianOptimization
from tensorflow.keras.layers import Dense, Dropout, BatchNormalization

# For additional metrics
from imblearn.metrics import geometric_mean_score, specificity_score
from sklearn.metrics import confusion_matrix
#import tensorflow_addons as tfa  # For focal loss function
import time
import matplotlib.pyplot as plt
import statistics
import seaborn as sns

# For storing results
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl


# For stratified Cross Validation
from sklearn.model_selection import RepeatedStratifiedKFold

# For notification
import chime

date = datetime.today().strftime('%m%d%y')  # For labelling purposes
from NeuralNetworkBase import NN


def weighted_loss_persample(weights, batchSize):
    def loss(y_true, y_pred):
        # The masks for the true and false values

        idx_1 = y_true == 1.
        idx_0 = y_true == 0.

        pred_1 = tf.boolean_mask(y_pred, idx_1)
        pred_1 = tf.expand_dims(pred_1, 1)

        true_1 = tf.boolean_mask(y_true, idx_1)
        true_1 = tf.expand_dims(true_1, 1)

        pred_0 = tf.boolean_mask(y_pred, idx_0)
        pred_0 = tf.expand_dims(pred_0, 1)

        true_0 = tf.boolean_mask(y_true, idx_0)
        true_0 = tf.expand_dims(true_0, 1)

        bce = tf.keras.losses.BinaryCrossentropy(reduction=tf.keras.losses.Reduction.NONE)

        # The losses for the 0 classes
        losses_0 = bce(pred_0, true_0)
        losses_0 = tf.reduce_sum(losses_0, 0)  # Add back up
        losses_0 = losses_0 * weights[0]

        # The losses for the 1 classes
        losses_1 = bce(pred_1, true_1)
        losses_1 = tf.reduce_sum(losses_1, 0)  # Add back up
        losses_1 = losses_1 * weights[1]

        # Add them back up and divide by batch size
        sum = losses_0 + losses_1
        total = sum / batchSize
        return total

    return loss

def age_encoderTX(data):
    age_map = {'04': 1, '05': 1, '06': 1,
               '07': 2, '08': 2, '09': 3,
               '10': 3, '11': 4, '12': 4, '13': 4}

    data['PAT_AGE'] = data['PAT_AGE'].map(age_map)

    data = pd.get_dummies(data, prefix_sep="__", dummy_na=True,
                          columns=['PAT_AGE'])

    data.loc[data["PAT_AGE__nan"] == 1, data.columns.str.startswith("PAT_AGE__")] = np.nan

    data = data.drop(['PAT_AGE__nan'], axis=1)

    data.rename(columns={'PAT_AGE__1.0': 'Ages 10-19',
                         'PAT_AGE__3.0': 'Ages 30-39',
                         'PAT_AGE__2.0': 'Ages 20-29',
                         'PAT_AGE__4.0': 'Ages 40+'}, inplace=True)

    return data


def age_encoderOK(data):
    age_map = {'10-14': 1, '15-19': 1, '20-24': 2,
               '25-29': 2, '30-34': 3, '35-39': 3,
               '40-44': 4, '45-49': 4, '50-54': 4}

    data['Age'] = data['Age'].map(age_map)

    data = pd.get_dummies(data, prefix_sep="__", dummy_na=True,
                          columns=['Age'])

    data.loc[data["Age__nan"] == 1, data.columns.str.startswith("Age__")] = np.nan

    data = data.drop(['Age__nan'], axis=1)

    data.rename(columns={'Age__1.0': 'Ages 10-19',
                         'Age__3.0': 'Ages 30-39',
                         'Age__2.0': 'Ages 20-29',
                         'Age__4.0': 'Ages 40+'}, inplace=True)

    return data


class fullNN(NN):

    def cleanDataTX(self, age):

        self.age = age

        quarter1 = pd.read_csv("Data/Texas_PUDF/PUDF_base1_1q2013_tab.txt", delimiter="\t", usecols=['RECORD_ID',
                                                                                                     'DISCHARGE',
                                                                                                     'SOURCE_OF_ADMISSION',
                                                                                                     'PAT_STATUS',
                                                                                                     'PAT_STATE',
                                                                                                     'COUNTY',
                                                                                                     'SEX_CODE',
                                                                                                     'RACE',
                                                                                                     'ETHNICITY',
                                                                                                     'PAT_AGE',
                                                                                                     'FIRST_PAYMENT_SRC',
                                                                                                     'SECONDARY_PAYMENT_SRC',
                                                                                                     'LENGTH_OF_STAY',
                                                                                                     'ADMITTING_DIAGNOSIS',
                                                                                                     'PRINC_DIAG_CODE',
                                                                                                     'OTH_DIAG_CODE_1',
                                                                                                     'OTH_DIAG_CODE_2',
                                                                                                     'OTH_DIAG_CODE_3',
                                                                                                     'OTH_DIAG_CODE_4',
                                                                                                     'OTH_DIAG_CODE_5',
                                                                                                     'OTH_DIAG_CODE_6',
                                                                                                     'OTH_DIAG_CODE_7',
                                                                                                     'OTH_DIAG_CODE_8',
                                                                                                     'OTH_DIAG_CODE_9',
                                                                                                     'OTH_DIAG_CODE_10',
                                                                                                     'OTH_DIAG_CODE_11',
                                                                                                     'OTH_DIAG_CODE_12',
                                                                                                     'OTH_DIAG_CODE_13',
                                                                                                     'OTH_DIAG_CODE_14',
                                                                                                     'OTH_DIAG_CODE_15',
                                                                                                     'OTH_DIAG_CODE_16',
                                                                                                     'OTH_DIAG_CODE_17',
                                                                                                     'OTH_DIAG_CODE_18',
                                                                                                     'OTH_DIAG_CODE_19',
                                                                                                     'OTH_DIAG_CODE_20',
                                                                                                     'OTH_DIAG_CODE_21',
                                                                                                     'OTH_DIAG_CODE_22',
                                                                                                     'OTH_DIAG_CODE_23',
                                                                                                     'OTH_DIAG_CODE_24', ],
                               dtype={'RECORD_ID': object,
                                      'DISCHARGE': object,
                                      'PAT_STATUS': object,
                                      'PAT_STATE': str,
                                      'COUNTY': str,
                                      'SOURCE_OF_ADMISSION': object,
                                      'SEX_CODE': object,
                                      'RACE': object,
                                      'ETHNICITY': object,
                                      'PAT_AGE': object,
                                      'FIRST_PAYMENT_SRC': object,
                                      'SECONDARY_PAYMENT_SRC': object,
                                      'LENGTH_OF_STAY': float,
                                      'ADMITTING_DIAGNOSIS': str,
                                      'PRINC_DIAG_CODE': str,
                                      'OTH_DIAG_CODE_1': str,
                                      'OTH_DIAG_CODE_2': str,
                                      'OTH_DIAG_CODE_3': str,
                                      'OTH_DIAG_CODE_4': str,
                                      'OTH_DIAG_CODE_5': str,
                                      'OTH_DIAG_CODE_6': str,
                                      'OTH_DIAG_CODE_7': str,
                                      'OTH_DIAG_CODE_8': str,
                                      'OTH_DIAG_CODE_9': str,
                                      'OTH_DIAG_CODE_10': str,
                                      'OTH_DIAG_CODE_11': str,
                                      'OTH_DIAG_CODE_12': str,
                                      'OTH_DIAG_CODE_13': str,
                                      'OTH_DIAG_CODE_14': str,
                                      'OTH_DIAG_CODE_15': str,
                                      'OTH_DIAG_CODE_16': str,
                                      'OTH_DIAG_CODE_17': str,
                                      'OTH_DIAG_CODE_18': str,
                                      'OTH_DIAG_CODE_19': str,
                                      'OTH_DIAG_CODE_20': str,
                                      'OTH_DIAG_CODE_21': str,
                                      'OTH_DIAG_CODE_22': str,
                                      'OTH_DIAG_CODE_23': str,
                                      'OTH_DIAG_CODE_24': str})

        quarter2 = pd.read_csv("Data/Texas_PUDF/PUDF_base1_2q2013_tab.txt", delimiter="\t", usecols=['RECORD_ID',
                                                                                                     'DISCHARGE',
                                                                                                     'SOURCE_OF_ADMISSION',
                                                                                                     'PAT_STATUS',
                                                                                                     'PAT_STATE',
                                                                                                     'COUNTY',
                                                                                                     'SEX_CODE',
                                                                                                     'RACE',
                                                                                                     'ETHNICITY',
                                                                                                     'PAT_AGE',
                                                                                                     'FIRST_PAYMENT_SRC',
                                                                                                     'SECONDARY_PAYMENT_SRC',
                                                                                                     'LENGTH_OF_STAY',
                                                                                                     'ADMITTING_DIAGNOSIS',
                                                                                                     'PRINC_DIAG_CODE',
                                                                                                     'OTH_DIAG_CODE_1',
                                                                                                     'OTH_DIAG_CODE_2',
                                                                                                     'OTH_DIAG_CODE_3',
                                                                                                     'OTH_DIAG_CODE_4',
                                                                                                     'OTH_DIAG_CODE_5',
                                                                                                     'OTH_DIAG_CODE_6',
                                                                                                     'OTH_DIAG_CODE_7',
                                                                                                     'OTH_DIAG_CODE_8',
                                                                                                     'OTH_DIAG_CODE_9',
                                                                                                     'OTH_DIAG_CODE_10',
                                                                                                     'OTH_DIAG_CODE_11',
                                                                                                     'OTH_DIAG_CODE_12',
                                                                                                     'OTH_DIAG_CODE_13',
                                                                                                     'OTH_DIAG_CODE_14',
                                                                                                     'OTH_DIAG_CODE_15',
                                                                                                     'OTH_DIAG_CODE_16',
                                                                                                     'OTH_DIAG_CODE_17',
                                                                                                     'OTH_DIAG_CODE_18',
                                                                                                     'OTH_DIAG_CODE_19',
                                                                                                     'OTH_DIAG_CODE_20',
                                                                                                     'OTH_DIAG_CODE_21',
                                                                                                     'OTH_DIAG_CODE_22',
                                                                                                     'OTH_DIAG_CODE_23',
                                                                                                     'OTH_DIAG_CODE_24', ],
                               dtype={'RECORD_ID': object,
                                      'DISCHARGE': object,
                                      'SOURCE_OF_ADMISSION': object,
                                      'PAT_STATUS': object,
                                      'PAT_STATE': str,
                                      'COUNTY': str,
                                      'SEX_CODE': object,
                                      'RACE': object,
                                      'ETHNICITY': object,
                                      'PAT_AGE': object,
                                      'FIRST_PAYMENT_SRC': object,
                                      'SECONDARY_PAYMENT_SRC': object,
                                      'LENGTH_OF_STAY': float,
                                      'ADMITTING_DIAGNOSIS': str,
                                      'PRINC_DIAG_CODE': str,
                                      'OTH_DIAG_CODE_1': str,
                                      'OTH_DIAG_CODE_2': str,
                                      'OTH_DIAG_CODE_3': str,
                                      'OTH_DIAG_CODE_4': str,
                                      'OTH_DIAG_CODE_5': str,
                                      'OTH_DIAG_CODE_6': str,
                                      'OTH_DIAG_CODE_7': str,
                                      'OTH_DIAG_CODE_8': str,
                                      'OTH_DIAG_CODE_9': str,
                                      'OTH_DIAG_CODE_10': str,
                                      'OTH_DIAG_CODE_11': str,
                                      'OTH_DIAG_CODE_12': str,
                                      'OTH_DIAG_CODE_13': str,
                                      'OTH_DIAG_CODE_14': str,
                                      'OTH_DIAG_CODE_15': str,
                                      'OTH_DIAG_CODE_16': str,
                                      'OTH_DIAG_CODE_17': str,
                                      'OTH_DIAG_CODE_18': str,
                                      'OTH_DIAG_CODE_19': str,
                                      'OTH_DIAG_CODE_20': str,
                                      'OTH_DIAG_CODE_21': str,
                                      'OTH_DIAG_CODE_22': str,
                                      'OTH_DIAG_CODE_23': str,
                                      'OTH_DIAG_CODE_24': str})

        quarter3 = pd.read_csv("Data/Texas_PUDF/PUDF_base1_3q2013_tab.txt", delimiter="\t", usecols=['RECORD_ID',
                                                                                                     'DISCHARGE',
                                                                                                     'SOURCE_OF_ADMISSION',
                                                                                                     'PAT_STATUS',
                                                                                                     'PAT_STATE',
                                                                                                     'COUNTY',
                                                                                                     'SEX_CODE',
                                                                                                     'RACE',
                                                                                                     'ETHNICITY',
                                                                                                     'PAT_AGE',
                                                                                                     'FIRST_PAYMENT_SRC',
                                                                                                     'SECONDARY_PAYMENT_SRC',
                                                                                                     'LENGTH_OF_STAY',
                                                                                                     'ADMITTING_DIAGNOSIS',
                                                                                                     'PRINC_DIAG_CODE',
                                                                                                     'OTH_DIAG_CODE_1',
                                                                                                     'OTH_DIAG_CODE_2',
                                                                                                     'OTH_DIAG_CODE_3',
                                                                                                     'OTH_DIAG_CODE_4',
                                                                                                     'OTH_DIAG_CODE_5',
                                                                                                     'OTH_DIAG_CODE_6',
                                                                                                     'OTH_DIAG_CODE_7',
                                                                                                     'OTH_DIAG_CODE_8',
                                                                                                     'OTH_DIAG_CODE_9',
                                                                                                     'OTH_DIAG_CODE_10',
                                                                                                     'OTH_DIAG_CODE_11',
                                                                                                     'OTH_DIAG_CODE_12',
                                                                                                     'OTH_DIAG_CODE_13',
                                                                                                     'OTH_DIAG_CODE_14',
                                                                                                     'OTH_DIAG_CODE_15',
                                                                                                     'OTH_DIAG_CODE_16',
                                                                                                     'OTH_DIAG_CODE_17',
                                                                                                     'OTH_DIAG_CODE_18',
                                                                                                     'OTH_DIAG_CODE_19',
                                                                                                     'OTH_DIAG_CODE_20',
                                                                                                     'OTH_DIAG_CODE_21',
                                                                                                     'OTH_DIAG_CODE_22',
                                                                                                     'OTH_DIAG_CODE_23',
                                                                                                     'OTH_DIAG_CODE_24', ],
                               dtype={'RECORD_ID': object,
                                      'DISCHARGE': object,
                                      'SOURCE_OF_ADMISSION': object,
                                      'PAT_STATUS': object,
                                      'PAT_STATE': str,
                                      'COUNTY': str,
                                      'SEX_CODE': object,
                                      'RACE': object,
                                      'ETHNICITY': object,
                                      'PAT_AGE': object,
                                      'FIRST_PAYMENT_SRC': object,
                                      'SECONDARY_PAYMENT_SRC': object,
                                      'LENGTH_OF_STAY': float,
                                      'ADMITTING_DIAGNOSIS': str,
                                      'PRINC_DIAG_CODE': str,
                                      'OTH_DIAG_CODE_1': str,
                                      'OTH_DIAG_CODE_2': str,
                                      'OTH_DIAG_CODE_3': str,
                                      'OTH_DIAG_CODE_4': str,
                                      'OTH_DIAG_CODE_5': str,
                                      'OTH_DIAG_CODE_6': str,
                                      'OTH_DIAG_CODE_7': str,
                                      'OTH_DIAG_CODE_8': str,
                                      'OTH_DIAG_CODE_9': str,
                                      'OTH_DIAG_CODE_10': str,
                                      'OTH_DIAG_CODE_11': str,
                                      'OTH_DIAG_CODE_12': str,
                                      'OTH_DIAG_CODE_13': str,
                                      'OTH_DIAG_CODE_14': str,
                                      'OTH_DIAG_CODE_15': str,
                                      'OTH_DIAG_CODE_16': str,
                                      'OTH_DIAG_CODE_17': str,
                                      'OTH_DIAG_CODE_18': str,
                                      'OTH_DIAG_CODE_19': str,
                                      'OTH_DIAG_CODE_20': str,
                                      'OTH_DIAG_CODE_21': str,
                                      'OTH_DIAG_CODE_22': str,
                                      'OTH_DIAG_CODE_23': str,
                                      'OTH_DIAG_CODE_24': str})

        quarter4 = pd.read_csv("Data/Texas_PUDF/PUDF_base1_4q2013_tab.txt", delimiter="\t", usecols=['RECORD_ID',
                                                                                                     'DISCHARGE',
                                                                                                     'SOURCE_OF_ADMISSION',
                                                                                                     'PAT_STATUS',
                                                                                                     'PAT_STATE',
                                                                                                     'COUNTY',
                                                                                                     'SEX_CODE',
                                                                                                     'RACE',
                                                                                                     'ETHNICITY',
                                                                                                     'PAT_AGE',
                                                                                                     'FIRST_PAYMENT_SRC',
                                                                                                     'SECONDARY_PAYMENT_SRC',
                                                                                                     'LENGTH_OF_STAY',
                                                                                                     'ADMITTING_DIAGNOSIS',
                                                                                                     'PRINC_DIAG_CODE',
                                                                                                     'OTH_DIAG_CODE_1',
                                                                                                     'OTH_DIAG_CODE_2',
                                                                                                     'OTH_DIAG_CODE_3',
                                                                                                     'OTH_DIAG_CODE_4',
                                                                                                     'OTH_DIAG_CODE_5',
                                                                                                     'OTH_DIAG_CODE_6',
                                                                                                     'OTH_DIAG_CODE_7',
                                                                                                     'OTH_DIAG_CODE_8',
                                                                                                     'OTH_DIAG_CODE_9',
                                                                                                     'OTH_DIAG_CODE_10',
                                                                                                     'OTH_DIAG_CODE_11',
                                                                                                     'OTH_DIAG_CODE_12',
                                                                                                     'OTH_DIAG_CODE_13',
                                                                                                     'OTH_DIAG_CODE_14',
                                                                                                     'OTH_DIAG_CODE_15',
                                                                                                     'OTH_DIAG_CODE_16',
                                                                                                     'OTH_DIAG_CODE_17',
                                                                                                     'OTH_DIAG_CODE_18',
                                                                                                     'OTH_DIAG_CODE_19',
                                                                                                     'OTH_DIAG_CODE_20',
                                                                                                     'OTH_DIAG_CODE_21',
                                                                                                     'OTH_DIAG_CODE_22',
                                                                                                     'OTH_DIAG_CODE_23',
                                                                                                     'OTH_DIAG_CODE_24', ],
                               dtype={'RECORD_ID': object,
                                      'DISCHARGE': object,
                                      'SOURCE_OF_ADMISSION': object,
                                      'PAT_STATUS': object,
                                      'PAT_STATE': str,
                                      'COUNTY': str,
                                      'SEX_CODE': object,
                                      'RACE': object,
                                      'ETHNICITY': object,
                                      'PAT_AGE': object,
                                      'FIRST_PAYMENT_SRC': object,
                                      'SECONDARY_PAYMENT_SRC': object,
                                      'LENGTH_OF_STAY': float,
                                      'ADMITTING_DIAGNOSIS': str,
                                      'PRINC_DIAG_CODE': str,
                                      'OTH_DIAG_CODE_1': str,
                                      'OTH_DIAG_CODE_2': str,
                                      'OTH_DIAG_CODE_3': str,
                                      'OTH_DIAG_CODE_4': str,
                                      'OTH_DIAG_CODE_5': str,
                                      'OTH_DIAG_CODE_6': str,
                                      'OTH_DIAG_CODE_7': str,
                                      'OTH_DIAG_CODE_8': str,
                                      'OTH_DIAG_CODE_9': str,
                                      'OTH_DIAG_CODE_10': str,
                                      'OTH_DIAG_CODE_11': str,
                                      'OTH_DIAG_CODE_12': str,
                                      'OTH_DIAG_CODE_13': str,
                                      'OTH_DIAG_CODE_14': str,
                                      'OTH_DIAG_CODE_15': str,
                                      'OTH_DIAG_CODE_16': str,
                                      'OTH_DIAG_CODE_17': str,
                                      'OTH_DIAG_CODE_18': str,
                                      'OTH_DIAG_CODE_19': str,
                                      'OTH_DIAG_CODE_20': str,
                                      'OTH_DIAG_CODE_21': str,
                                      'OTH_DIAG_CODE_22': str,
                                      'OTH_DIAG_CODE_23': str,
                                      'OTH_DIAG_CODE_24': str})

        # Combining all the quarters into one dataframe
        frames = [quarter1, quarter2, quarter3, quarter4]
        year2013 = pd.concat(frames)

        # Insurance Codes
        medicare = ['16', 'MA', 'MB']
        medicaid = ['MC']
        sc = ['09', 'ZZ']
        other = ['10', '11', 'AM', 'CI', 'LI',
                 'LM', '12', '13', '14', '15',
                 'BL', 'CH', 'HM', 'OF', 'WC',
                 'DS', 'VA', 'TV']

        # County Information ----------------------------------------------------------------------------

        # fips code of each county in Border Area
        on_border = ['043', '047', '061', '105',
                     '109', '127', '131', '137',
                     '141', '163', '215', '229',
                     '243', '247', '261', '271',
                     '283', '323', '311', '371',
                     '377', '385', '389', '427',
                     '435', '443', '463', '465',
                     '479', '489', '505', '507']

        year2013['On Border'] = 0
        year2013.loc[year2013['COUNTY'].isin(on_border), ['On Border']] = 1
        # If the value is missing in county, set new value to missing
        year2013['On Border'].loc[year2013['COUNTY'].isna()] = np.nan

        # Filter by hospital delivery
        # Checks if a delivery-type code (V27*) is in any of the icd9 columns
        year2013 = year2013.loc[(year2013['ADMITTING_DIAGNOSIS'].str.startswith('V27'))
                                | year2013['PRINC_DIAG_CODE'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_1'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_2'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_3'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_4'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_5'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_6'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_7'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_8'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_9'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_10'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_11'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_12'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_13'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_14'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_15'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_16'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_17'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_18'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_19'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_20'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_21'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_22'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_23'].str.startswith('V27')
                                | year2013['OTH_DIAG_CODE_24'].str.startswith('V27')]

        # Drop year2012
        year2013 = year2013[year2013['DISCHARGE'] != '2012Q4']

        # Selecting appropriate age groups
        year2013 = year2013.query('PAT_AGE >= "04" & PAT_AGE <= "13"')

        if age == 'Ordinal':

            # Encode variables
            enc = OrdinalEncoder()
            year2013[["PAT_AGE"]] = enc.fit_transform(year2013[["PAT_AGE"]])

        elif age == 'Categorical':
            year2013 = age_encoderTX(year2013)

        # Drop Invalid gender rows
        year2013 = year2013[year2013['SEX_CODE'] != 'U']

        # Re-label Invalid gender rows
        year2013['SEX_CODE'] = year2013['SEX_CODE'].replace('M', 'F')

        # Replace all tick marks with nan, either impute or drop later
        year2013 = year2013.replace('`', np.NaN)

        # Replace sex code nan with F, since at this point we should have only females in the df
        year2013['SEX_CODE'] = year2013['SEX_CODE'].replace(np.NaN, 'F')

        # Changes payment sources from codes to corresponding categories
        year2013['FIRST_PAYMENT_SRC'] = year2013['FIRST_PAYMENT_SRC'].replace(medicare, "Medicare")
        year2013['FIRST_PAYMENT_SRC'] = year2013['FIRST_PAYMENT_SRC'].replace(medicaid, "Medicaid")
        year2013['FIRST_PAYMENT_SRC'] = year2013['FIRST_PAYMENT_SRC'].replace(sc,
                                                                              "Self-pay or Charity")
        year2013['FIRST_PAYMENT_SRC'] = year2013['FIRST_PAYMENT_SRC'].replace(other,
                                                                              "Other Insurance")

        year2013['SECONDARY_PAYMENT_SRC'] = year2013['SECONDARY_PAYMENT_SRC'].replace(medicare,
                                                                                      "Medicare")
        year2013['SECONDARY_PAYMENT_SRC'] = year2013['SECONDARY_PAYMENT_SRC'].replace(medicaid,
                                                                                      "Medicaid")
        year2013['SECONDARY_PAYMENT_SRC'] = year2013['SECONDARY_PAYMENT_SRC'].replace(sc,
                                                                                      "Self-pay or Charity")
        year2013['SECONDARY_PAYMENT_SRC'] = year2013['SECONDARY_PAYMENT_SRC'].replace(other,
                                                                                      "Other Insurance")

        year2013 = pd.get_dummies(year2013, dummy_na=False,
                                  columns=['FIRST_PAYMENT_SRC', 'SECONDARY_PAYMENT_SRC'])

        # Create category columns
        year2013['Medicaid'] = 0
        year2013['Medicare'] = 0
        year2013['Self-pay or Charity'] = 0
        year2013['Other Insurance'] = 0

        year2013['Medicaid'] = np.where(year2013['FIRST_PAYMENT_SRC_Medicaid'] == 1, 1,
                                        year2013[
                                            'Medicaid'])  # Change to 1 if 1, otherwise leave as is
        year2013['Medicaid'] = np.where(year2013['SECONDARY_PAYMENT_SRC_Medicaid'] == 1, 1,
                                        year2013['Medicaid'])
        year2013['Medicare'] = np.where(year2013['FIRST_PAYMENT_SRC_Medicare'] == 1, 1,
                                        year2013['Medicare'])
        year2013['Medicare'] = np.where(year2013['SECONDARY_PAYMENT_SRC_Medicare'] == 1, 1,
                                        year2013['Medicare'])
        year2013['Self-pay or Charity'] = np.where(
            year2013['FIRST_PAYMENT_SRC_Self-pay or Charity'] == 1, 1,
            year2013['Self-pay or Charity'])
        year2013['Self-pay or Charity'] = np.where(
            year2013['SECONDARY_PAYMENT_SRC_Self-pay or Charity'] == 1, 1,
            year2013['Self-pay or Charity'])
        year2013['Other Insurance'] = np.where(year2013['FIRST_PAYMENT_SRC_Other Insurance'] == 1, 1,
                                               year2013['Other Insurance'])
        year2013['Other Insurance'] = np.where(year2013['SECONDARY_PAYMENT_SRC_Other Insurance'] == 1,
                                               1, year2013['Other Insurance'])

        # Drop columns with dummies
        year2013.drop(columns=['FIRST_PAYMENT_SRC_Medicaid',
                               'SECONDARY_PAYMENT_SRC_Medicaid',
                               'FIRST_PAYMENT_SRC_Medicare',
                               'SECONDARY_PAYMENT_SRC_Medicare',
                               'FIRST_PAYMENT_SRC_Self-pay or Charity',
                               'SECONDARY_PAYMENT_SRC_Self-pay or Charity',
                               'FIRST_PAYMENT_SRC_Other Insurance',
                               'SECONDARY_PAYMENT_SRC_Other Insurance']
                      , axis=1, inplace=True)

        # Rename Race columns
        """
        year2013['RACE'] = year2013['RACE'].replace('1', 'Native American')
        year2013['RACE'] = year2013['RACE'].replace('2', 'Asian or Pacific Islander')
        year2013['RACE'] = year2013['RACE'].replace('3', 'Black')
        year2013['RACE'] = year2013['RACE'].replace('4', 'White')
        year2013['RACE'] = year2013['RACE'].replace('5', 'Other Race')
        """

        # Columns for scanning ICD9 codes
        diagnosisColumns = ['ADMITTING_DIAGNOSIS',
                            'PRINC_DIAG_CODE',
                            'OTH_DIAG_CODE_1',
                            'OTH_DIAG_CODE_2',
                            'OTH_DIAG_CODE_3',
                            'OTH_DIAG_CODE_4',
                            'OTH_DIAG_CODE_5',
                            'OTH_DIAG_CODE_6',
                            'OTH_DIAG_CODE_7',
                            'OTH_DIAG_CODE_8',
                            'OTH_DIAG_CODE_9',
                            'OTH_DIAG_CODE_10',
                            'OTH_DIAG_CODE_11',
                            'OTH_DIAG_CODE_12',
                            'OTH_DIAG_CODE_13',
                            'OTH_DIAG_CODE_14',
                            'OTH_DIAG_CODE_15',
                            'OTH_DIAG_CODE_16',
                            'OTH_DIAG_CODE_17',
                            'OTH_DIAG_CODE_18',
                            'OTH_DIAG_CODE_19',
                            'OTH_DIAG_CODE_20',
                            'OTH_DIAG_CODE_21',
                            'OTH_DIAG_CODE_22',
                            'OTH_DIAG_CODE_23',
                            'OTH_DIAG_CODE_24']

        # Creating a dictionary to hold keys and values
        diseaseDictionary = {}

        diseaseDictionary['Obesity'] = ['V853', 'V854', '27800', '27801', '27803', '6491']
        # diseaseDictionary['Pregnancy resulting from assisted reproductive technology'] = ['V2385']
        diseaseDictionary['Cocaine dependence'] = ['3042', '3056']
        diseaseDictionary['Amphetamine dependence'] = ['3044', '3057']
        diseaseDictionary['Gestational diabetes mellitus'] = ['6488']
        diseaseDictionary['Pre-existing diabetes mellitus'] = ['250', '6480']
        diseaseDictionary['Anxiety'] = ['3000']
        diseaseDictionary['Anemia NOS'] = ['2859']
        diseaseDictionary['Iron deficiency anemia'] = ['280']
        diseaseDictionary['Other anemia'] = ['281']
        diseaseDictionary['Depression'] = ['311']
        diseaseDictionary['Primigravidas at the extremes of maternal age'] = ['6595', 'V2381', 'V2383']
        diseaseDictionary['Hemorrhagic disorders due to intrinsic circulating antibodies'] = ['2865']
        diseaseDictionary['Systemic lupus erythematosus'] = ['7100']
        diseaseDictionary['Lupus erythematosus'] = ['6954']
        diseaseDictionary['Autoimmune disease not elsewhere classified'] = ['27949']
        diseaseDictionary['Pure hypercholesterolemia'] = ['2720']
        diseaseDictionary['Unspecified vitamin D deficiency'] = ['2689']
        diseaseDictionary['Proteinuria'] = ['7910']
        diseaseDictionary['Tobacco use disorder'] = ['3051', '6490']
        diseaseDictionary['History of tobacco use'] = ['V1582']
        diseaseDictionary['Hypertension'] = ['401']
        diseaseDictionary['Hypertensive heart disease'] = ['402']
        diseaseDictionary['Chronic venous hypertension'] = ['4593']
        diseaseDictionary['Unspecified renal disease in pregnancy without mention of hypertension'] = ['6462']
        diseaseDictionary['Chronic kidney disease'] = ['585']
        diseaseDictionary['Hypertensive kidney disease'] = ['403']
        diseaseDictionary['Hypertensive heart and chronic kidney disease'] = ['404']
        diseaseDictionary['Renal failure not elsewhere classified'] = ['586']
        diseaseDictionary['Infections of genitourinary tract in pregnancy'] = ['6466']
        diseaseDictionary['UTI'] = ['5990']
        diseaseDictionary['Personal history of trophoblastic disease'] = ['V131']
        diseaseDictionary['Supervision of high-risk pregnancy with history of trophoblastic disease'] = ['V231']
        diseaseDictionary['Thrombophilia'] = ['28981']
        diseaseDictionary['History of premature delivery'] = ['V1321']
        diseaseDictionary['Hemorrhage in early pregnancy'] = ['640']
        diseaseDictionary[
            'Congenital abnormalities of the uterus including those complicating pregnancy, childbirth, or the puerperium'] = [
            '6540', '7522', '7523']
        # diseaseDictionary['Multiple Gestations'] = ['651']
        diseaseDictionary['Fetal Growth Restriction'] = ['764']
        diseaseDictionary['Asthma'] = ['493']
        diseaseDictionary['Obstructive Sleep Apnea'] = ['32723']
        diseaseDictionary['Other cardiovascular diseases complicating pregnancy and childbirth or the puerperium'] = [
            '6486']
        diseaseDictionary['Sickle cell disease'] = ['28260']
        diseaseDictionary['Thyroid Disease'] = ['240', '241', '242', '243', '244', '245', '246']
        diseaseDictionary['Inadequate Prenatal Care'] = ['V237']
        diseaseDictionary['Periodontal disease'] = ['523']
        diseaseDictionary['Preeclampsia/Eclampsia'] = ['6424', '6425', '6426', '6427']

        # Adds Disease column
        for disease in diseaseDictionary:
            year2013[disease] = 0  # This is how to add columns and default to 0

        # Filling out the diseases
        for disease in diseaseDictionary:
            for codes in diseaseDictionary[disease]:
                for col in diagnosisColumns:
                    year2013.loc[year2013[col].str.startswith(codes, na=False), [disease]] = 1

        eclampExclude = ['64243', '64253', '64263', '64273']  # Exclude codes ending in 3
        for codes in eclampExclude:
            for col in diagnosisColumns:
                year2013.loc[
                    year2013[col].str.startswith(codes, na=False), ['Preeclampsia/Eclampsia']] = 0

        # Drop columns with ICD-9 codes
        year2013.drop(columns=
                      ['ADMITTING_DIAGNOSIS', 'PRINC_DIAG_CODE', 'OTH_DIAG_CODE_1',
                       'OTH_DIAG_CODE_2', 'OTH_DIAG_CODE_3', 'OTH_DIAG_CODE_4',
                       'OTH_DIAG_CODE_5', 'OTH_DIAG_CODE_6', 'OTH_DIAG_CODE_7',
                       'OTH_DIAG_CODE_8', 'OTH_DIAG_CODE_9', 'OTH_DIAG_CODE_10',
                       'OTH_DIAG_CODE_11', 'OTH_DIAG_CODE_12', 'OTH_DIAG_CODE_13',
                       'OTH_DIAG_CODE_14', 'OTH_DIAG_CODE_15', 'OTH_DIAG_CODE_16',
                       'OTH_DIAG_CODE_17', 'OTH_DIAG_CODE_18', 'OTH_DIAG_CODE_19',
                       'OTH_DIAG_CODE_20', 'OTH_DIAG_CODE_21', 'OTH_DIAG_CODE_22',
                       'OTH_DIAG_CODE_23', 'OTH_DIAG_CODE_24'], axis=1, inplace=True)

        # Drop the columns that will not be used
        year2013.drop(
            columns=['LENGTH_OF_STAY', 'SOURCE_OF_ADMISSION', 'RECORD_ID', 'SEX_CODE', 'COUNTY', 'PAT_STATE',
                     'PAT_STATUS'], axis=1, inplace=True)

        # year2013 = (year2013.loc[(year2013['Pregnancy resulting from assisted reproductive technology'] == 0)])
        year2013 = (year2013.loc[(year2013['Multiple Gestations'] == 0)])

        # Setting dummies to true makes a column for each category that states whether or not it is missing (0 or 1).
        year2013 = pd.get_dummies(year2013, prefix_sep="__", dummy_na=True,
                                  columns=['DISCHARGE', 'ETHNICITY'])

        # Propogates the missing values via the indicator columns
        year2013.loc[year2013["DISCHARGE__nan"] == 1, year2013.columns.str.startswith("DISCHARGE__")] = np.nan
        year2013.loc[year2013["ETHNICITY__nan"] == 1, year2013.columns.str.startswith("ETHNICITY__")] = np.nan

        # Drops the missingness indicator columns
        year2013 = year2013.drop(['DISCHARGE__nan'], axis=1)
        year2013 = year2013.drop(['ETHNICITY__nan'], axis=1)

        """
        year2013.rename(columns={'ETHNICITY__1': 'Hispanic', 'ETHNICITY__2': 'Non-Hispanic'},
                        inplace=True)
        """

        African_Am = year2013.loc[year2013['RACE'] == '3']
        African_Am.drop(columns=['RACE'], inplace=True)
        African_Am.to_csv('Data/AfricanAmerican_' + date + '.csv', index=False)

        Native_Am = year2013.loc[year2013['RACE'] == '1']
        Native_Am.drop(columns=['RACE'], inplace=True)
        Native_Am.to_csv('Data/NativeAmerican_' + date + '.csv', index=False)

        # One hot encoding race
        year2013 = pd.get_dummies(year2013, prefix_sep="__", dummy_na=True,
                                  columns=['RACE'])
        year2013.loc[
            year2013["RACE__nan"] == 1, year2013.columns.str.startswith("RACE__")] = np.nan

        year2013 = year2013.drop(['RACE__nan'], axis=1)

        # Create new combined race and ethnicity columns
        year2013['White Hispanic'] = 0
        year2013['Black Hispanic'] = 0
        year2013['White Non-Hispanic'] = 0
        year2013['Black Non-Hispanic'] = 0
        year2013['Asian/Pacific Islander Hispanic'] = 0
        year2013['American Indian/Eskimo/Aleut Hispanic'] = 0
        year2013['Asian/Pacific Islander Non-Hispanic'] = 0
        year2013['American Indian/Eskimo/Aleut Non-Hispanic'] = 0
        year2013['Other Race Hispanic'] = 0
        year2013['Other Race Non-Hispanic'] = 0

        # Fill out columns with appropriate numbers
        year2013['White Hispanic'] = np.where(((year2013['RACE__4'] == 1) & (year2013['ETHNICITY__1'] == 1)), 1,
                                              year2013['White Hispanic'])
        year2013['Black Hispanic'] = np.where(((year2013['RACE__3'] == 1) & (year2013['ETHNICITY__1'] == 1)), 1,
                                              year2013['Black Hispanic'])
        year2013['Asian/Pacific Islander Hispanic'] = np.where(
            ((year2013['RACE__2'] == 1) & (year2013['ETHNICITY__1'] == 1)), 1,
            year2013['Asian/Pacific Islander Hispanic'])
        year2013['American Indian/Eskimo/Aleut Hispanic'] = np.where(
            ((year2013['RACE__1'] == 1) & (year2013['ETHNICITY__1'] == 1)), 1,
            year2013['American Indian/Eskimo/Aleut Hispanic'])
        year2013['Other Race Hispanic'] = np.where(((year2013['RACE__5'] == 1) & (year2013['ETHNICITY__1'] == 1)),
                                                   1, year2013['Other Race Hispanic'])
        year2013['White Non-Hispanic'] = np.where(((year2013['RACE__4'] == 1) & (year2013['ETHNICITY__2'] == 1)), 1,
                                                  year2013['White Non-Hispanic'])
        year2013['Black Non-Hispanic'] = np.where(((year2013['RACE__3'] == 1) & (year2013['ETHNICITY__2'] == 1)), 1,
                                                  year2013['Black Non-Hispanic'])
        year2013['Asian/Pacific Islander Non-Hispanic'] = np.where(
            ((year2013['RACE__2'] == 1) & (year2013['ETHNICITY__2'] == 1)), 1,
            year2013['Asian/Pacific Islander Non-Hispanic'])
        year2013['American Indian/Eskimo/Aleut Non-Hispanic'] = np.where(
            ((year2013['RACE__1'] == 1) & (year2013['ETHNICITY__2'] == 1)), 1,
            year2013['American Indian/Eskimo/Aleut Non-Hispanic'])
        year2013['Other Race Non-Hispanic'] = np.where(
            ((year2013['RACE__5'] == 1) & (year2013['ETHNICITY__2'] == 1)), 1, year2013['Other Race Non-Hispanic'])

        # Drop original race and ethnicity columns
        year2013.drop(columns=['RACE__1', 'RACE__2', 'RACE__3', 'RACE__4',
                               'RACE__5', 'ETHNICITY__1', 'ETHNICITY__2'], axis=1, inplace=True)

        year2013.to_csv('Data/year2013_' + date + '.csv', index=False)

        return year2013, African_Am, Native_Am

    def cleanDataOK(self, dropMetro, age='Ordinal'):

        self.age = age

        self.dropMetro = dropMetro

        ok2017 = pd.read_csv(
            r"file:///home/rachel/Documents/Preeclampsia_Research/Data/Oklahom_PUDF_2020.08.27/2017%20IP/pudf_cd.txt",
            sep=",")
        ok2018 = pd.read_csv(
            r"file:///home/rachel/Documents/Preeclampsia_Research/Data/Oklahom_PUDF_2020.08.27/2018%20IP/pudf_cdv2.txt",
            sep=",")

        # Dropping unneeded columns
        ok2017.drop(columns=['pk_pudf', 'id_hups', 'cd_hospital_type', 'cd_admission_type_src', 'no_total_chgs',
                             'cd_drg_hci', 'cd_mdc', 'cd_ecode_cause_1',
                             'cd_ecode_cause_2', 'cd_ecode_cause_3'], inplace=True)
        ok2018.drop(columns=['pk_pudf', 'id_hups', 'cd_hospital_type', 'cd_admission_type_src', 'no_total_chgs',
                             'cd_drg_hci', 'cd_mdc', 'cd_ecode_cause_1',
                             'cd_ecode_cause_2', 'cd_ecode_cause_3'], inplace=True)

        ok2017.columns = ['State', 'Zip', 'County', 'Sex', 'Race', 'Marital_status', 'Age', 'admit_year',
                          'admit_month', 'admit_day',
                          'discharge_year', 'discharge_month', 'discharge_day', 'Length_of_stay', 'Status',
                          'Insurance', 'pdx', 'dx1', 'dx2', 'dx3',
                          'dx4', 'dx5', 'dx6', 'dx7', 'dx8', 'dx9', 'dx10', 'dx11', 'dx12', 'dx13', 'dx14', 'dx15',
                          'ppoa', 'poa1', 'poa2', 'poa3', 'poa4', 'poa5',
                          'poa6', 'poa7', 'poa8', 'poa9', 'poa10', 'poa11', 'poa12', 'poa13',
                          'poa14', 'poa15', 'ppx', 'px1', 'px2', 'px3', 'px4', 'px5', 'px6',
                          'px7', 'px8', 'px9', 'px10', 'px11', 'px12', 'px13', 'px14', 'px15']

        ok2018.columns = ['State', 'Zip', 'County', 'Sex', 'Race', 'Marital_status', 'Age', 'admit_year',
                          'admit_month', 'admit_day',
                          'discharge_year', 'discharge_month', 'discharge_day', 'Length_of_stay', 'Status',
                          'Insurance', 'pdx', 'dx1', 'dx2', 'dx3',
                          'dx4', 'dx5', 'dx6', 'dx7', 'dx8', 'dx9', 'dx10', 'dx11', 'dx12', 'dx13', 'dx14', 'dx15',
                          'ppoa', 'poa1', 'poa2', 'poa3', 'poa4', 'poa5',
                          'poa6', 'poa7', 'poa8', 'poa9', 'poa10', 'poa11', 'poa12', 'poa13',
                          'poa14', 'poa15', 'ppx', 'px1', 'px2', 'px3', 'px4', 'px5', 'px6',
                          'px7', 'px8', 'px9', 'px10', 'px11', 'px12', 'px13', 'px14', 'px15']

        ok2017 = (ok2017.loc[(ok2017['pdx'].str.startswith('Z37'))
                             | ok2017['dx1'].str.startswith('Z37')
                             | ok2017['dx2'].str.startswith('Z37')
                             | ok2017['dx3'].str.startswith('Z37')
                             | ok2017['dx4'].str.startswith('Z37')
                             | ok2017['dx5'].str.startswith('Z37')
                             | ok2017['dx6'].str.startswith('Z37')
                             | ok2017['dx8'].str.startswith('Z37')
                             | ok2017['dx9'].str.startswith('Z37')
                             | ok2017['dx10'].str.startswith('Z37')
                             | ok2017['dx11'].str.startswith('Z37')
                             | ok2017['dx12'].str.startswith('Z37')
                             | ok2017['dx13'].str.startswith('Z37')
                             | ok2017['dx14'].str.startswith('Z37')
                             | ok2017['dx15'].str.startswith('Z37')])

        ok2018 = (ok2018.loc[(ok2018['pdx'].str.startswith('Z37'))
                             | ok2018['dx1'].str.startswith('Z37')
                             | ok2018['dx2'].str.startswith('Z37')
                             | ok2018['dx3'].str.startswith('Z37')
                             | ok2018['dx4'].str.startswith('Z37')
                             | ok2018['dx5'].str.startswith('Z37')
                             | ok2018['dx6'].str.startswith('Z37')
                             | ok2018['dx8'].str.startswith('Z37')
                             | ok2018['dx9'].str.startswith('Z37')
                             | ok2018['dx10'].str.startswith('Z37')
                             | ok2018['dx11'].str.startswith('Z37')
                             | ok2018['dx12'].str.startswith('Z37')
                             | ok2018['dx13'].str.startswith('Z37')
                             | ok2018['dx14'].str.startswith('Z37')
                             | ok2018['dx15'].str.startswith('Z37')])

        # Fix missing values
        ok2017['State'] = np.where(ok2017['State'] == '99', np.NaN, ok2017['State'])
        ok2018['State'] = np.where(ok2018['State'] == '99', np.NaN, ok2018['State'])

        ok2017['Zip'] = np.where(ok2017['Zip'] == 99999.0, np.NaN, ok2017['Zip'])
        ok2018['Zip'] = np.where(ok2018['Zip'] == 99999.0, np.NaN, ok2018['Zip'])

        ok2017['Marital_status'] = np.where(ok2017['Marital_status'] == 'U', np.NaN, ok2017['Marital_status'])
        ok2018['Marital_status'] = np.where(ok2018['Marital_status'] == 'U', np.NaN, ok2018['Marital_status'])

        ok2017['Sex'] = np.where(ok2017['Sex'] == 'U', np.NaN, ok2017['Sex'])
        ok2018['Sex'] = np.where(ok2018['Sex'] == 'U', np.NaN, ok2018['Sex'])

        ok2017['Age'] = np.where(ok2017['Age'] == '99', np.NaN, ok2017['Age'])
        ok2018['Age'] = np.where(ok2018['Age'] == '99', np.NaN, ok2018['Age'])

        ok2017['Status'] = np.where(ok2017['Status'] == '99', np.NaN, ok2017['Status'])
        ok2018['Status'] = np.where(ok2018['Status'] == '99', np.NaN, ok2018['Status'])

        # Creating Insurance Binary Columns
        ok2017['Medicaid'] = 0
        ok2017['Medicare'] = 0
        ok2017['Self-pay'] = 0
        ok2017['Other Insurance'] = 0

        # Creating Insurance Binary Columns
        ok2018['Medicaid'] = 0
        ok2018['Medicare'] = 0
        ok2018['Self-pay'] = 0
        ok2018['Other Insurance'] = 0

        # Filling out appropriate Columns
        ok2017['Medicaid'] = np.where(ok2017['Insurance'] == 3, 1,
                                      ok2017['Medicaid'])  # Change to 1 if 1, otherwise leave as is
        ok2017['Medicare'] = np.where(ok2017['Insurance'] == 2, 1, ok2017['Medicare'])
        ok2017['Self-pay'] = np.where(ok2017['Insurance'] == 6, 1, ok2017['Self-pay'])
        ok2017['Other Insurance'] = np.where(ok2017['Insurance'].isin([1, 4, 5, 7]), int(1),
                                             ok2017['Other Insurance'])

        # For Missing Values, 9 is unknown in their dictionary
        ok2017['Medicaid'] = np.where(ok2017['Insurance'] == 9, np.NaN,
                                      ok2017['Medicaid'])  # Change to 1 if 1, otherwise leave as is
        ok2017['Medicare'] = np.where(ok2017['Insurance'] == 9, np.NaN, ok2017['Medicare'])
        ok2017['Self-pay'] = np.where(ok2017['Insurance'] == 9, np.NaN, ok2017['Self-pay'])
        ok2017['Other Insurance'] = np.where(ok2017['Insurance'] == 9, np.NaN, ok2017['Other Insurance'])

        # Filling out appropriate Columns
        ok2018['Medicaid'] = np.where(ok2018['Insurance'] == 3, 1,
                                      ok2018['Medicaid'])  # Change to 1 if 1, otherwise leave as is
        ok2018['Medicare'] = np.where(ok2018['Insurance'] == 2, 1, ok2018['Medicare'])
        ok2018['Self-pay'] = np.where(ok2018['Insurance'] == 6, 1, ok2018['Self-pay'])
        ok2018['Other Insurance'] = np.where(ok2018['Insurance'].isin([1, 4, 5, 7]), int(1),
                                             ok2018['Other Insurance'])

        # For Missing Values, 9 is unkown in their dictionary
        ok2018['Medicaid'] = np.where(ok2018['Insurance'] == 9, np.NaN,
                                      ok2018['Medicaid'])  # Change to 1 if 1, otherwise leave as is
        ok2018['Medicare'] = np.where(ok2018['Insurance'] == 9, np.NaN, ok2018['Medicare'])
        ok2018['Self-pay'] = np.where(ok2018['Insurance'] == 9, np.NaN, ok2018['Self-pay'])
        ok2018['Other Insurance'] = np.where(ok2018['Insurance'] == 9, np.NaN, ok2018['Other Insurance'])

        # Fixing incorrect values
        ok2018['Medicaid'] = np.where(ok2018['Insurance'].isin([11, 14]), np.NaN,
                                      ok2018['Medicaid'])  # Change to 1 if 1, otherwise leave as is
        ok2018['Medicare'] = np.where(ok2018['Insurance'].isin([11, 14]), np.NaN, ok2018['Medicare'])
        ok2018['Self-pay'] = np.where(ok2018['Insurance'].isin([11, 14]), np.NaN, ok2018['Self-pay'])
        ok2018['Other Insurance'] = np.where(ok2018['Insurance'].isin([11, 14]), np.NaN, ok2018['Other Insurance'])

        # Dropping Insurance column
        ok2017.drop(columns=['Insurance'], inplace=True)
        ok2018.drop(columns=['Insurance'], inplace=True)

        # Re-label Invalid gender rows
        ok2017['Sex'] = ok2017['Sex'].replace('M', 'F')
        ok2018['Sex'] = ok2018['Sex'].replace('M', 'F')

        # Selecting appropriate age groups
        ok2017 = ok2017.query('Age >= "01" & Age <= "50-54" | Age == "99"')
        ok2018 = ok2018.query('Age >= "01" & Age <= "50-54" | Age == "99"')

        if age == 'Ordinal':

            # Ordinal Encode Age
            enc = OrdinalEncoder()
            ok2017[["Age"]] = enc.fit_transform(ok2017[["Age"]])
            ok2018[["Age"]] = enc.fit_transform(ok2018[["Age"]])

        elif age == 'Categorical':
            ok2017 = age_encoderOK(ok2017)
            ok2018 = age_encoderOK(ok2018)

        # Re-label Race
        ok2017['Race'].replace('W', 'White', inplace=True)
        ok2017['Race'].replace('B', 'Black', inplace=True)
        ok2017['Race'].replace('I', 'Native American', inplace=True)
        ok2017['Race'].replace('O', 'Other/Unknown', inplace=True)

        ok2018['Race'].replace('W', 'White', inplace=True)
        ok2018['Race'].replace('B', 'Black', inplace=True)
        ok2018['Race'].replace('I', 'Native American', inplace=True)
        ok2018['Race'].replace('O', 'Other/Unknown', inplace=True)

        # Read in list of Counties and their designation
        urbanRural = pd.read_csv(
            r"file:///home/rachel/Documents/Preeclampsia_Research/Data/County%20Metropolitan%20Classification.csv")
        urbanRural['county name'] = urbanRural['county name'].str.replace(' County', '')
        urbanRural['Metro status'] = urbanRural['Metro status'].replace('Metropolitan', 'Urban')
        urbanRural['Metro status'] = urbanRural['Metro status'].replace('Nonmetropolitan', 'Rural')
        urbanRural.drop(columns='value', inplace=True)

        # Match capitalization
        ok2017['County'] = ok2017.County.str.capitalize()
        ok2018['County'] = ok2018.County.str.capitalize()

        # Join to include whether county is urban or rural
        ok2017 = ok2017.merge(urbanRural, left_on=['County', 'State'], right_on=['county name', 'State'],
                              how='left')
        ok2018 = ok2018.merge(urbanRural, left_on=['County', 'State'], right_on=['county name', 'State'],
                              how='left')

        # Keeping admit month as proxy for whether when they developed preeclampsia
        ok2017.drop(columns=['admit_year', 'admit_day', 'discharge_month', 'discharge_year', 'discharge_day'],
                    inplace=True)
        ok2018.drop(columns=['admit_year', 'admit_day', 'discharge_month', 'discharge_year', 'discharge_day'],
                    inplace=True)

        # Re-label marriage status
        ok2017['Marital_status'] = ok2017['Marital_status'].replace('M', 'Married')
        ok2018['Marital_status'] = ok2018['Marital_status'].replace('M', 'Married')
        ok2017['Marital_status'] = ok2017['Marital_status'].replace('N', 'Unmarried')
        ok2018['Marital_status'] = ok2018['Marital_status'].replace('N', 'Unmarried')

        # A list of relevant columns
        diagnosisColumns = ['pdx', 'dx1', 'dx2', 'dx3',
                            'dx4', 'dx5', 'dx6', 'dx7',
                            'dx8', 'dx9', 'dx10', 'dx11',
                            'dx12', 'dx13', 'dx14', 'dx15']

        # Creating a dictionary to hold keys and values
        diseaseDictionary = {}

        diseaseDictionary['Obesity'] = ['E66', 'O9921', 'O9981', 'O9984', 'Z683', 'Z684', 'Z713', 'Z9884']
        diseaseDictionary['Pregnancy resulting from assisted reproductive technology'] = ['O0981']
        diseaseDictionary['Cocaine dependence'] = ['F14', 'T405']
        diseaseDictionary['Amphetamine dependence'] = ['F15', 'F19', 'P044', 'T4362']
        diseaseDictionary['Gestational diabetes mellitus'] = ['O244', 'P700']
        diseaseDictionary['Pre-existing diabetes mellitus'] = ['E10', 'E11', 'O240', 'O241', 'O243', 'O248', 'O249']
        diseaseDictionary['Anxiety'] = ['F064', 'F41']
        diseaseDictionary['Anemia NOS'] = ['D51']
        diseaseDictionary['Iron deficiency anemia'] = ['D50']
        diseaseDictionary['Other anemia'] = ['D64', 'D59', 'D489', 'D53', 'O990']
        diseaseDictionary['Depression'] = ['F32', 'F341', 'F33', 'F0631', 'Z139', 'Z1331', 'Z1332']
        diseaseDictionary['Primigravidas at the extremes of maternal age'] = ['O095', 'O096']
        diseaseDictionary['Hemorrhagic disorders due to intrinsic circulating antibodies'] = ['D683']
        diseaseDictionary['Systemic lupus erythematosus'] = ['M32']
        diseaseDictionary['Lupus erythematosus'] = ['L93', 'D6862']
        diseaseDictionary['Autoimmune disease not elsewhere classified'] = ['D89']
        diseaseDictionary['Pure hypercholesterolemia'] = ['E780']
        diseaseDictionary['Unspecified vitamin D deficiency'] = ['E55']
        diseaseDictionary['Proteinuria'] = ['D511', 'N06', 'O121', 'O122', 'R80']
        diseaseDictionary['Current Smoker'] = ['F172']
        diseaseDictionary['Hypertension'] = ['G932', 'I10', 'I14', 'I15', 'I272', 'I674', 'I973', 'O10', 'O13',
                                             'O16', 'R030']
        diseaseDictionary['Hypertensive heart disease'] = ['I11']
        diseaseDictionary['Chronic venous hypertension'] = ['I873']
        diseaseDictionary['Unspecified renal disease in pregnancy without mention of hypertension'] = ['O2683',
                                                                                                       'O9089']
        diseaseDictionary['Chronic kidney disease'] = ['D631', 'E0822', 'E0922', 'E0922', 'E1022', 'E1122', 'E1322',
                                                       'N18']
        diseaseDictionary['Hypertensive kidney disease'] = ['I12']
        diseaseDictionary['Hypertensive heart and chronic kidney disease'] = ['I13']
        diseaseDictionary['Renal failure not elsewhere classified'] = ['N19']
        diseaseDictionary['Infections of genitourinary tract in pregnancy'] = ['O23', 'O861', 'O862', 'O868']
        diseaseDictionary['UTI'] = ['O0338', 'O0388', 'O0488', 'O0788', 'O0883', 'N136', 'N390', 'N99521', 'N99531']
        diseaseDictionary['Personal history of trophoblastic disease'] = ['Z8759', 'O01']
        diseaseDictionary['Supervision of high-risk pregnancy with history of trophoblastic disease'] = ['O091']
        diseaseDictionary['Thrombophilia'] = ['D685', 'D686']
        diseaseDictionary['History of premature delivery'] = ['Z8751']
        diseaseDictionary['Hemorrhage in early pregnancy'] = ['O20']
        diseaseDictionary[
            'Congenital abnormalities of the uterus including those complicating pregnancy, childbirth, or the puerperium'] = [
            'O34', 'O340']
        diseaseDictionary['Multiple Gestations'] = ['O30']
        diseaseDictionary['Inadequate Prenatal Care'] = ['O093']
        diseaseDictionary['Periodontal disease'] = ['E08630', 'E09630', 'E10630', 'E11630', '13630', 'K05', 'K06',
                                                    'K08129']
        diseaseDictionary['Intrauterine Death'] = ['O364']
        diseaseDictionary['Preeclampsia/Eclampsia'] = ['O14', 'O15']

        # New Additions
        """
        diseaseDictionary['Edema'] = ['R609']
        diseaseDictionary['Hyperreflexia'] = ['R292']
        diseaseDictionary['Oliguria'] = ['R34']
        diseaseDictionary['Headache'] = ['R41']
        diseaseDictionary['Vomiting'] = ['R1110']
        """

        # Adds Disease column
        for disease in diseaseDictionary:
            ok2017[disease] = 0  # This is how to add columns and default to 0

            # Adds Disease column
        for disease in diseaseDictionary:
            ok2018[disease] = 0  # This is how to add columns and default to 0

        # Filling out the diseases
        for disease in diseaseDictionary:
            for codes in diseaseDictionary[disease]:
                for col in diagnosisColumns:
                    ok2017.loc[ok2017[col].str.startswith(codes, na=False), [disease]] = 1

        for disease in diseaseDictionary:
            for codes in diseaseDictionary[disease]:
                for col in diagnosisColumns:
                    ok2018.loc[ok2018[col].str.startswith(codes, na=False), [disease]] = 1

        ok2017.drop(columns=['State', 'Zip', 'Sex', 'County', 'Length_of_stay', 'Status', 'pdx',
                             'dx1', 'dx2', 'dx3', 'dx4', 'dx5', 'dx6', 'dx7',
                             'dx8', 'dx9', 'dx10', 'dx11', 'dx12', 'dx13', 'dx14',
                             'dx15', 'ppoa', 'poa1', 'poa2', 'poa3', 'poa4',
                             'poa5', 'poa6', 'poa7', 'poa8', 'poa9', 'poa10',
                             'poa11', 'poa12', 'poa13', 'poa14', 'poa15', 'ppx',
                             'px1', 'px2', 'px3', 'px4', 'px5', 'px6', 'px7',
                             'px8', 'px9', 'px10', 'px11', 'px12', 'px13', 'px14',
                             'px15', 'county name'], inplace=True)

        ok2018.drop(columns=['State', 'Zip', 'Sex', 'County', 'Length_of_stay', 'Status', 'pdx',
                             'dx1', 'dx2', 'dx3', 'dx4', 'dx5', 'dx6', 'dx7',
                             'dx8', 'dx9', 'dx10', 'dx11', 'dx12', 'dx13', 'dx14',
                             'dx15', 'ppoa', 'poa1', 'poa2', 'poa3', 'poa4',
                             'poa5', 'poa6', 'poa7', 'poa8', 'poa9', 'poa10',
                             'poa11', 'poa12', 'poa13', 'poa14', 'poa15', 'ppx',
                             'px1', 'px2', 'px3', 'px4', 'px5', 'px6', 'px7',
                             'px8', 'px9', 'px10', 'px11', 'px12', 'px13', 'px14',
                             'px15', 'county name'], inplace=True)

        # ok2017 = (ok2017.loc[(ok2017['Multiple Gestations'] == 0)])
        # ok2018 = (ok2018.loc[(ok2018['Multiple Gestations'] == 0)])

        ok2017 = (ok2017.loc[(ok2017['Pregnancy resulting from assisted reproductive technology'] == 0)])
        ok2018 = (ok2018.loc[(ok2018['Pregnancy resulting from assisted reproductive technology'] == 0)])

        # Setting dummies to true makes a column for each category that states whether or not it is missing (0 or 1).
        ok2017 = pd.get_dummies(ok2017, prefix_sep="__", dummy_na=True,
                                columns=['Race', 'Marital_status', 'Metro status'])

        # Propogates the missing values via the indicator columns
        ok2017.loc[ok2017["Race__nan"] == 1, ok2017.columns.str.startswith("Race__")] = np.nan
        ok2017.loc[ok2017["Marital_status__nan"] == 1, ok2017.columns.str.startswith("Marital_status__")] = np.nan
        ok2017.loc[ok2017["Metro status__nan"] == 1, ok2017.columns.str.startswith("Metro status__")] = np.nan

        # Drops the missingness indicator columns
        ok2017 = ok2017.drop(['Race__nan'], axis=1)
        ok2017 = ok2017.drop(['Marital_status__nan'], axis=1)
        ok2017 = ok2017.drop(['Metro status__nan'], axis=1)

        # Setting dummies to true makes a column for each category that states whether or not it is missing (0 or 1).
        ok2018 = pd.get_dummies(ok2018, prefix_sep="__", dummy_na=True,
                                columns=['Race', 'Marital_status', 'Metro status'])

        # Propogates the missing values via the indicator columns
        ok2018.loc[ok2018["Race__nan"] == 1, ok2018.columns.str.startswith("Race__")] = np.nan
        ok2018.loc[ok2018["Marital_status__nan"] == 1, ok2018.columns.str.startswith("Marital_status__")] = np.nan
        ok2018.loc[ok2018["Metro status__nan"] == 1, ok2018.columns.str.startswith("Metro status__")] = np.nan

        # Drops the missingness indicator columns
        ok2018 = ok2018.drop(['Race__nan'], axis=1)
        ok2018 = ok2018.drop(['Marital_status__nan'], axis=1)
        ok2018 = ok2018.drop(['Metro status__nan'], axis=1)

        ok2017.rename(columns={'Race__White': 'White',
                               'Race__Native American': 'Native American',
                               'Race__Black': 'Black',
                               'Race__Other/Unknown': 'Other/Unknown Race',
                               'Marital_status__Married': 'Married',
                               'Marital_status__Unmarried': 'Unmarried',
                               'Metro status__Rural': 'Rural',
                               'Metro status__Urban': 'Urban'}, inplace=True)

        ok2018.rename(columns={'Race__White': 'White',
                               'Race__Native American': 'Native American',
                               'Race__Black': 'Black',
                               'Race__Other/Unknown': 'Other/Unknown Race',
                               'Marital_status__Married': 'Married',
                               'Marital_status__Unmarried': 'Unmarried',
                               'Metro status__Rural': 'Rural',
                               'Metro status__Urban': 'Urban'}, inplace=True)

        if (dropMetro == True):
            ok2017.drop(columns=['Rural', 'Urban'], inplace=True)
            ok2018.drop(columns=['Rural', 'Urban'], inplace=True)

        # ok2017.to_csv('Data/Oklahoma_Clean/ok2017_Incomplete.csv', index=False)
        # ok2018.to_csv('Data/Oklahoma_Clean/ok2018_Incomplete.csv', index=False)

        return ok2017, ok2018

    def prepData(self, age, data):

        self.age = age

        data = pd.read_csv(data)

        X = data.drop(columns='Preeclampsia/Eclampsia')
        Y = data['Preeclampsia/Eclampsia']


        return X, Y

    def setData(self, X_train, X_test, Y_train, Y_test):

        self.X_train = X_train
        self.X_test = X_test
        self.Y_train = Y_train
        self.Y_test = Y_test

    def featureSelection(self, numFeatures, method):

        self.method = method

        if method == 1:
            model = XGBClassifier()
            model.fit(self.X_train, self.Y_train)

            # Save graph
            ax = plot_importance(model, max_num_features=numFeatures)
            fig1 = pyplot.gcf()
            #pyplot.show()
            fig1.savefig(self.dataset + 'XGBoostTopFeatures.png', bbox_inches='tight')

            # Get and save best features
            feature_important = model.get_booster().get_fscore()
            keys = list(feature_important.keys())
            values = list(feature_important.values())

            data = pd.DataFrame(data=values, index=keys, columns=["score"]).sort_values(by="score", ascending=False)
            XGBoostFeatures = list(data.index[0:numFeatures])
            return XGBoostFeatures

        if method == 2:
            # instantiate SelectKBest to determine 20 best features
            fs = SelectKBest(score_func=chi2, k=numFeatures)
            fs.fit(self.X_train, self.Y_train)
            df_scores = pd.DataFrame(fs.scores_)
            df_columns = pd.DataFrame(self.X_train.columns)
            # concatenate dataframes
            feature_scores = pd.concat([df_columns, df_scores], axis=1)
            feature_scores.columns = ['Feature_Name', 'Chi2 Score']  # name output columns
            feature_scores.sort_values(by=['Chi2 Score'], ascending=False, inplace=True)
            features = feature_scores.iloc[0:numFeatures]
            chi2Features = features['Feature_Name']
            self.Chi2features = features
            return chi2Features

        if method == 3:
            # Mutual Information features
            fs = SelectKBest(score_func=mutual_info_classif, k=numFeatures)
            fs.fit(self.X_train, self.Y_train)
            df_scores = pd.DataFrame(fs.scores_)
            df_columns = pd.DataFrame(self.X_train.columns)
            # concatenate dataframes
            feature_scores = pd.concat([df_columns, df_scores], axis=1)
            feature_scores.columns = ['Feature_Name', 'MI Score']  # name output columns
            feature_scores.sort_values(by=['MI Score'], ascending=False, inplace=True)
            features = feature_scores.iloc[0:numFeatures]
            mutualInfoFeatures = features['Feature_Name']
            self.MIFeatures = features
            return mutualInfoFeatures


class NoTune(fullNN):

    def buildModel(self, topFeatures, batchSize, initializer, epochs, alpha=None, gamma=None, biasInit=0):

        self.biasInit = biasInit
        self.start_time = time.time()
        self.best_hps = {'num_layers': 3,
                         'dense_activation_0': 'tanh',
                         'dense_activation_1': 'relu',
                         'dense_activation_2': 'relu',
                         'units_0': 30,
                         'units_1': 36,
                         'units_2': 45,
                         'final_activation': 'sigmoid',
                         'optimizer': 'Adam',
                         'learning_rate': 0.001}

        self.alpha = alpha
        self.gamma = gamma
        self.epochs = epochs
        LOG_DIR = f"{int(time.time())}"

        # Set all to numpy arrays
        self.X_train = self.X_train[topFeatures].to_numpy()
        self.Y_train = self.Y_train.to_numpy()
        self.X_test = self.X_test[topFeatures].to_numpy()
        self.Y_test = self.Y_test.to_numpy()

        inputSize = self.X_train.shape[1]

        self.batch_size = batchSize

        self.training_generator = BalancedBatchGenerator(self.X_train, self.Y_train,
                                                         batch_size=self.batch_size,
                                                         sampler=RandomOverSampler(),
                                                         random_state=42)

        self.validation_generator = BalancedBatchGenerator(self.X_test, self.Y_test,
                                                           batch_size=self.batch_size,
                                                           sampler=RandomOverSampler(),
                                                           random_state=42)
        # define the keras model
        tf.keras.backend.clear_session()
        self.model = tf.keras.models.Sequential()
        self.model.add(tf.keras.Input(shape=(inputSize,)))

        # Hidden Layers
        for i in range(self.best_hps['num_layers']):
            self.model.add(
                Dense(units=self.best_hps['units_' + str(i)], activation=self.best_hps['dense_activation_' + str(i)]))
            # self.model.add(Dropout(0.20))
            # self.model.add(BatchNormalization(momentum=0.60))

        # Class weights
        class_weights = class_weight.compute_class_weight('balanced', np.unique(self.Y_train), self.Y_train)
        class_weight_dict = dict(enumerate(class_weights))
        pos = class_weight_dict[1]
        neg = class_weight_dict[0]

        bias = np.log(pos / neg)

        if biasInit == 0:
            # Final Layer
            self.model.add(Dense(1, activation=self.best_hps['final_activation']))
        elif biasInit == 1:
            # Final Layer
            self.model.add(Dense(1, activation=self.best_hps['final_activation'],
                                 bias_initializer=tf.keras.initializers.Constant(
                                     value=bias)))  # kernel_initializer=initializer,

        # Loss Function
        if alpha != None or gamma != None:
            loss = tfa.losses.SigmoidFocalCrossEntropy(alpha=alpha, gamma=gamma)
            self.loss = "focal_loss"
        else:
            loss = 'binary_crossentropy'
            self.loss = "binary-crossentropy"

        # Compilation
        self.model.compile(optimizer=tf.keras.optimizers.Adam(learning_rate=self.best_hps['learning_rate']),
                           loss=loss,
                           metrics=['accuracy',
                                    tf.keras.metrics.Precision(),
                                    tf.keras.metrics.Recall(),
                                    tf.keras.metrics.AUC()])

        self.history = self.model.fit(self.training_generator, epochs=epochs, validation_data=self.validation_generator,
                                      verbose=2)#, class_weight=class_weight_dict)

    def evaluateModel(self):
        """
        wb = Workbook()
        wsResults = wb.active
        wsResults.title = "Results"
        wsHyp = wb.create_sheet(title="Hyperparameters")
        wsFeatures = wb.create_sheet(title="Features")
        wsGraphs = wb.create_sheet(title="Graphs")
        filename = self.dataset + '_' + date

        # Graphing results
        plt.clf()
        plt.cla()
        plt.close()

        # plt.ylim(0.40, 0.66)
        plt.plot(self.history.history['auc'])
        plt.plot(self.history.history['val_auc'])
        plt.title('model auc')
        plt.ylabel('auc')
        plt.xlabel('epoch')
        plt.legend(['train', 'validation'], loc='upper right')
        figname = self.dataset + '_trainTestAUC' + date + '.png'
        plt.savefig(figname, bbox_inches="tight")
        img = openpyxl.drawing.image.Image(figname)
        img.anchor = 'A1'
        #wsGraphs.add_image(img)
        # plt.show()

        plt.clf()
        plt.cla()
        plt.close()

        # plt.ylim(0.0, 0.15)
        plt.plot(self.history.history['loss'])
        plt.plot(self.history.history['val_loss'])
        plt.title('model loss')
        plt.ylabel('loss')
        plt.xlabel('epoch')
        plt.legend(['train', 'validation'], loc='upper right')
        figname = self.dataset + '_trainTestLoss' + date + '.png'
        plt.savefig(figname, bbox_inches="tight")
        img = openpyxl.drawing.image.Image(figname)
        img.anchor = 'J1'
        #wsGraphs.add_image(img)
        # plt.show()

        # Feature Selection
        if self.method == 1:
            img = openpyxl.drawing.image.Image(self.dataset + 'XGBoostTopFeatures.png')
            img.anchor = 'A1'
            wsFeatures.add_image(img)

        elif self.method == 2:
            for r in dataframe_to_rows(self.Chi2features, index=False, header=True):
                wsFeatures.append(r)

        elif self.method == 3:
            for r in dataframe_to_rows(self.MIFeatures, index=False, header=True):
                wsFeatures.append(r)

        """
        y_predict = (self.model.predict(self.X_test) > 0.5).astype("int32")

        self.specificity = specificity_score(self.Y_test, y_predict)

        self.gmean = geometric_mean_score(self.Y_test, y_predict)

        score = self.model.evaluate(self.X_test, self.Y_test, verbose=0)
        self.loss = score[0]
        self.accuracy = score[1]
        self.AUC = score[4]
        self.predictedNo = y_predict.sum()
        self.trueNo = self.Y_test.sum()
        self.recall = score[3]
        self.precision = score[2]
        self.tn, self.fp, self.fn, self.tp = confusion_matrix(self.Y_test, y_predict).ravel()

        results = [['loss', self.loss],
                   ['accuracy', self.accuracy],
                   ['AUC', self.AUC],
                   ['specificity', self.specificity],
                   ['recall', self.recall],
                   ['precision', self.precision],
                   ['gmean', self.gmean],
                   ['True Positive', self.tp],
                   ['True Negative', self.tn],
                   ['False Positive', self.fp],
                   ['False Negative', self.fn]]

        df = pd.DataFrame(results, columns=['Metric', 'Value'])

        """
        for r in dataframe_to_rows(df, index=False, header=True):
            wsResults.append(r)
        """
        # df.to_csv(filename, index=False)

        print(f'Total Cases: {len(y_predict)}')
        print(f'Predict #: {y_predict.sum()} / True # {self.Y_test.sum()}')
        print(f'True Positives #: {self.tp} / True Negatives # {self.tn}')
        print(f'False Positives #: {self.fp} / False Negatives # {self.fn}')
        print(f'Test loss: {score[0]:.6f} / Test accuracy: {score[1]:.6f} / Test AUC: {score[4]:.6f}')
        print(f'Test Recall: {score[3]:.6f} / Test Precision: {score[2]:.6f}')
        print(f'Test Specificity: {self.specificity:.6f}')
        print(f'Test Gmean: {self.gmean:.6f}')

        mins = (time.time() - self.start_time) / 60  # Time in seconds
        """
        self.best_hps['Batch Size'] = self.batch_size
        self.best_hps['Age'] = self.age
        self.best_hps['epochs'] = self.epochs
        self.best_hps['loss'] = self.loss
        self.best_hps['alpha'] = self.alpha
        self.best_hps['gamma'] = self.gamma
        self.best_hps['split1'] = self.split1
        self.best_hps['split2'] = self.split2
        self.best_hps['bias'] = self.biasInit
        self.best_hps['time(mins)'] = mins

        df = pd.DataFrame(data=self.best_hps, index=[0])
        df = (df.T)
        for r in dataframe_to_rows(df, index=True, header=False):
            wsHyp.append(r)

        wb.save(filename + '.xlsx')
        """
        return self.AUC, self.gmean, self.precision, self.recall, self.specificity, self.tp, self.fp, self.tn, self.fn, self.loss, self.history


class NoGen(NoTune):

    def buildModel(self, topFeatures, batchSize, initializer, epochs, alpha=None, gamma=None, biasInit=0):
        self.biasInit = biasInit
        self.start_time = time.time()
        self.best_hps = {'num_layers': 3,
                         'dense_activation_0': 'tanh',
                         'dense_activation_1': 'relu',
                         'dense_activation_2': 'relu',
                         'units_0': 30,
                         'units_1': 36,
                         'units_2': 45,
                         'final_activation': 'sigmoid',
                         'optimizer': 'Adam',
                         'learning_rate': 0.001}

        self.alpha = alpha
        self.gamma = gamma
        self.epochs = epochs
        LOG_DIR = f"{int(time.time())}"

        # Set all to numpy arrays
        self.X_train = self.X_train[topFeatures]
        self.Y_train = self.Y_train
        self.X_val = self.X_val[topFeatures]
        self.Y_val = self.Y_val
        self.X_test = self.X_test[topFeatures]
        self.Y_test = self.Y_test

        inputSize = self.X_train.shape[1]

        self.batch_size = batchSize

        # define the keras model
        tf.keras.backend.clear_session()
        self.model = tf.keras.models.Sequential()
        self.model.add(tf.keras.Input(shape=(inputSize,)))

        # Hidden Layers
        for i in range(self.best_hps['num_layers']):
            self.model.add(
                Dense(units=self.best_hps['units_' + str(i)], activation=self.best_hps['dense_activation_' + str(i)]))
            self.model.add(Dropout(0.20))
            self.model.add(BatchNormalization(momentum=0.60))

        # Class weights
        class_weights = class_weight.compute_class_weight('balanced', np.unique(self.Y_train), self.Y_train)
        class_weight_dict = dict(enumerate(class_weights))
        pos = class_weight_dict[1]
        neg = class_weight_dict[0]

        class_weight_dict[0] = 1 / self.Y_train.value_counts()[0]
        class_weight_dict[1] = 1 / self.Y_train.value_counts()[1]

        # class_weight_dict[0] = (self.Y_train.value_counts()[1]/self.Y_train.value_counts()[0])
        # class_weight_dict[1] = 1

        class_weight_dict[0] = 1
        class_weight_dict[1] = 5

        bias = np.log(pos / neg)

        if biasInit == 0:
            # Final Layer
            self.model.add(Dense(1, activation=self.best_hps['final_activation']))
        elif biasInit == 1:
            # Final Layer
            self.model.add(Dense(1, activation=self.best_hps['final_activation'],
                                 bias_initializer=tf.keras.initializers.Constant(
                                     value=bias)))  # kernel_initializer=initializer,

        # Loss Function
        if alpha != None or gamma != None:
            loss = tfa.losses.SigmoidFocalCrossEntropy(alpha=alpha, gamma=gamma)
            self.loss = "focal_loss"
        else:
            loss = 'binary_crossentropy'
            self.loss = "binary-crossentropy"

        # Compilation
        self.model.compile(optimizer=tf.keras.optimizers.Adam(learning_rate=self.best_hps['learning_rate']),
                           loss=weighted_loss_persample(class_weight_dict, batchSize),
                           metrics=['accuracy',
                                    tf.keras.metrics.Precision(),
                                    tf.keras.metrics.Recall(),
                                    tf.keras.metrics.AUC()])

        self.history = self.model.fit(self.X_train, self.Y_train, batch_size=batchSize,
                                      epochs=epochs, validation_data=(self.X_val, self.Y_val),
                                      verbose=2)


if __name__ == "__main__":

    rskf = RepeatedStratifiedKFold(n_splits=2, n_repeats=2, random_state=36851234)

    filename = 'Results/Oklahoma/Full/NoTune/CVTests/Test_binary2-2cv_noreg'

    modelWeighted = NoTune(filename)

    X, y = modelWeighted.prepData(age='Categorical',
                           data='Data/Processed/Oklahoma/Complete/Full/Outliers/Chi2_Categorical.csv')

    aucList = []
    gmeanList = []
    precisionList = []
    recallList = []
    specList = []
    lossList = []
    historyList = []

    for train_index, test_index in rskf.split(X, y):
        X_train, X_test = X.iloc[train_index, :], X.iloc[test_index, :]
        y_train, y_test = y.iloc[train_index], y.iloc[test_index]

        modelWeighted.setData(X_train, X_test, y_train, y_test)

        features = modelWeighted.featureSelection(numFeatures=20, method=2)
        # For hand-tuning
        modelWeighted.buildModel(features, batchSize=4096, initializer='RandomUniform',
                                 biasInit=1, epochs=100)

        auc, gmean, precision, recall, specificity, tp, fp, tn, fn, loss, history = modelWeighted.evaluateModel()

        aucList.append(auc)
        gmeanList.append(gmean)
        precisionList.append(precision)
        recallList.append(recall)
        specList.append(specificity)
        lossList.append(loss)
        historyList.append(history) #List of lists, will each entry history of a particular run




    def saveResults(lossList, aucList, gmeanList, recallList, specList, precisionList, filename):

        lossMean = statistics.mean(lossList)
        aucMean = statistics.mean(aucList)
        gmeanMean = statistics.mean(gmeanList)
        specMean = statistics.mean(specList)
        recallMean = statistics.mean(recallList)
        precMean = statistics.mean(precisionList)

        wb = Workbook()
        wsMean = wb.active
        wsMean.title = "Mean Results"
        wsResults= wb.create_sheet(title="Results Full")

        # Save Mean results
        results = [['loss', lossMean],
                   ['AUC',  aucMean],
                   ['gmean',  gmeanMean],
                   ['specificity', specMean],
                   ['recall', recallMean],
                   ['precision', precMean]]

        df = pd.DataFrame(results, columns=['Metric', 'Mean'])

        for r in dataframe_to_rows(df, index=False, header=True):
            wsMean.append(r)

        # Save the whole lists of each cross-validation

        df = pd.DataFrame({'Loss': lossList, 'AUC': aucList, 'Gmean':gmeanList, 'Recall': recallList,
                           'Precision': precisionList})

        for r in dataframe_to_rows(df, index=False, header=True):
            wsResults.append(r)

        wb.save(filename + '.xlsx')


    def plotAvg(historyList):
        aucAvg = []
        aucValAvg = []
        lossAvg = []
        lossValAvg = []

        for i in range(len(historyList[0].history['auc'])): # Iterate through each epoch
            # Clear list
            auc = []
            aucVal = []
            loss = []
            lossVal = []

            for j in range(len(historyList)): # Iterate through each history object

                # Append each model's measurement for epoch i
                auc.append(historyList[j].history['auc'][i])
                aucVal.append(historyList[j].history['val_auc'][i])
                loss.append(historyList[j].history['loss'][i])
                lossVal.append(historyList[j].history['val_loss'][i])

            # Once get measurement for each model, get average measurement for that epoch
            aucAvg.append(statistics.mean(auc))
            aucValAvg.append(statistics.mean(aucVal))
            lossAvg.append(statistics.mean(loss))
            lossValAvg.append(statistics.mean(lossVal))

        # Graphing results
        plt.clf()
        plt.cla()
        plt.close()

        # plt.ylim(0.40, 0.66)
        plt.plot(aucAvg)
        plt.plot(aucValAvg)
        plt.title('model auc')
        plt.ylabel('auc')
        plt.xlabel('epoch')
        plt.legend(['training', 'validation'], loc='upper right')
        figname = filename + '_trainTestAUC' + date + '.png'
        plt.savefig(figname, bbox_inches="tight")
        img = openpyxl.drawing.image.Image(figname)
        img.anchor = 'A1'
        wsGraphs.add_image(img)

        plt.clf()
        plt.cla()
        plt.close()

        # plt.ylim(0.40, 0.66)
        plt.plot(lossAvg)
        plt.plot(lossValAvg)
        plt.title('model loss')
        plt.ylabel('loss')
        plt.xlabel('epoch')
        plt.legend(['training', 'validation'], loc='upper right')
        figname = filename + '_trainTestLoss' + date + '.png'
        plt.savefig(figname, bbox_inches="tight")


    saveResults(lossList, aucList, gmeanList, recallList, specList, precisionList, filename)
    plotAvg(historyList)


    chime.success()