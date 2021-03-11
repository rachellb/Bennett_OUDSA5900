
#For handling data
import pandas as pd
import numpy as np
from datetime import datetime
import time

#For imputing data
from sklearn.experimental import enable_iterative_imputer
from sklearn.impute import IterativeImputer
from sklearn.utils import class_weight
#For feature selection
from sklearn.feature_selection import SelectKBest, chi2
from xgboost import XGBClassifier
from matplotlib import pyplot
from xgboost import plot_importance
from sklearn.model_selection import train_test_split
from sklearn.feature_selection import mutual_info_classif

#For Outlier Detection
from sklearn.ensemble import IsolationForest
from sklearn.covariance import EllipticEnvelope
from sklearn.neighbors import LocalOutlierFactor
from sklearn.svm import OneClassSVM

#For NN and tuning
import tensorflow as tf
tf.random.Generator = None  # Patch for a bug, Testing
#import kerastuner
#from kerastuner.tuners import Hyperband, BayesianOptimization, RandomSearch
from tensorflow.keras.layers import Dense, Dropout, BatchNormalization
import tensorflow_addons as tfa # For focal loss function
from tensorflow.keras.regularizers import l2

# For balancing batches
from imblearn.keras import BalancedBatchGenerator
from imblearn.over_sampling import RandomOverSampler

#For additional metrics
from imblearn.metrics import geometric_mean_score, specificity_score
from sklearn.metrics import confusion_matrix
import matplotlib.pyplot as plt
from openpyxl import Workbook  # For storing results
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import time
import os
date = datetime.today().strftime('%m%d%y') # For labelling purposes


def weighted_loss_persample(weights, batchSize):
    def loss(y_true, y_pred):

        # Returns the indices of the positive and negative samples
        idx_1 = y_true == 1.
        idx_0 = y_true == 0.

        # Returns only the elements that are in the positive class. Expand_dims ensures it is still a tensor shape
        pred_1 = tf.boolean_mask(y_pred, idx_1)
        pred_1 = tf.expand_dims(pred_1, 1)

        true_1 = tf.boolean_mask(y_true, idx_1)
        true_1 = tf.expand_dims(true_1, 1)

        # Returns only the elements that are in the negative class
        pred_0 = tf.boolean_mask(y_pred, idx_0)
        pred_0 = tf.expand_dims(pred_0, 1)

        true_0 = tf.boolean_mask(y_true, idx_0)
        true_0 = tf.expand_dims(true_0, 1)

        # By setting reduction = None, ensures that what is returned is an array-like of length (batch_size),
        # with each entry being the loss for an individual sample in the batch.
        bce = tf.keras.losses.BinaryCrossentropy(reduction=tf.keras.losses.Reduction.NONE)

        # The losses for the negative class, multiplies each loss by the associated weights
        losses_0 = bce(pred_0, true_0) * weights[0]

        # adds all the losses together, returns a scalar sum of losses
        losses_0 = tf.reduce_sum(losses_0, 0)


        # The losses for the positive class
        losses_1 = bce(pred_1, true_1) * weights[1]
        losses_1 = tf.reduce_sum(losses_1, 0)

        # Add them back up and divide by batch size
        sum = losses_0 + losses_1

        # The result is the normal binary crossentropy, with the losses added together and divided by batch sizes,
        # giving average loss per batch
        total = sum/batchSize
        return total

    return loss

class NN():

    def __init__(self, dataset):
        print("Welcome to class initializer...\n")

        self.dataset = dataset

    def prepData(self, age, data):

        self.age=age

        self.data = pd.read_csv(data)

    def imputeData(self, data1, data2=None):
        MI_Imp = IterativeImputer()  # Scikitlearn's Iterative imputer

        if (data2 is not None):  # If running both datasets
            if (data1.isnull().values.any() == True | data2.isnull().values.any() == True):
                data = data1.append(data2)
                self.data = pd.DataFrame(np.round(MI_Imp.fit_transform(data)), columns=data.columns)
            else:
                self.data = data1.append(data2)
        else:
            if (data1.isnull().values.any() == True):
                self.data = pd.DataFrame(np.round(MI_Imp.fit_transform(data1)), columns=data1.columns)
            else:
                self.data = data1

    def splitData(self, testSize, valSize):
        self.split1=5
        self.split2=107
        X = self.data.drop(columns='Preeclampsia/Eclampsia')
        Y = self.data['Preeclampsia/Eclampsia']
        self.X_train, self.X_test, self.Y_train, self.Y_test = train_test_split(X, Y, stratify=Y, test_size=testSize,
                                                                                random_state=self.split1)
        self.X_train, self.X_val, self.Y_train, self.Y_val = train_test_split(self.X_train, self.Y_train,
                                                                              stratify=self.Y_train, test_size=valSize,
                                                                              random_state=self.split2)

    def detectOutliers(self, method, con):

        print(self.X_train.shape, self.Y_train.shape)

        if method == 'iso':
            out = IsolationForest(contamination=con)
        elif method == 'lof':
            out = LocalOutlierFactor(contamination=con)
        elif method == 'ocsvm':
            out = OneClassSVM(nu=0.01)
        elif method == 'ee':
            out = EllipticEnvelope(contamination=con)

        yhat = out.fit_predict(self.X_train)

        # select all rows that are not outliers
        mask = (yhat != -1)

        self.X_train = self.X_train.loc[mask]
        self.Y_train = self.Y_train.loc[mask]

        print(self.X_train.shape, self.Y_train.shape)

    def featureSelection(self, numFeatures, method):

        self.method = method

        if method == 1:
            model = XGBClassifier()
            model.fit(self.X_train, self.Y_train)

            # Save graph
            ax = plot_importance(model, max_num_features=numFeatures)
            fig1 = pyplot.gcf()
            #pyplot.show()
            fig1.savefig(self.dataset + 'XGBoostTopFeatures.png', bbox_inches='tight')

            # Get and save best features
            feature_important = model.get_booster().get_fscore()
            keys = list(feature_important.keys())
            values = list(feature_important.values())

            data = pd.DataFrame(data=values, index=keys, columns=["score"]).sort_values(by="score", ascending=False)
            XGBoostFeatures = list(data.index[0:numFeatures])
            return XGBoostFeatures

        if method == 2:
            # instantiate SelectKBest to determine 20 best features
            fs = SelectKBest(score_func=chi2, k=numFeatures)
            fs.fit(self.X_train, self.Y_train)
            df_scores = pd.DataFrame(fs.scores_)
            df_columns = pd.DataFrame(self.X_train.columns)
            # concatenate dataframes
            feature_scores = pd.concat([df_columns, df_scores], axis=1)
            feature_scores.columns = ['Feature_Name', 'Chi2 Score']  # name output columns
            feature_scores.sort_values(by=['Chi2 Score'], ascending=False, inplace=True)
            features = feature_scores.iloc[0:numFeatures]
            chi2Features = features['Feature_Name']
            self.Chi2features = features
            return chi2Features

        if method == 3:
            # Mutual Information features
            fs = SelectKBest(score_func=mutual_info_classif, k=numFeatures)
            fs.fit(self.X_train, self.Y_train)
            df_scores = pd.DataFrame(fs.scores_)
            df_columns = pd.DataFrame(self.X_train.columns)
            # concatenate dataframes
            feature_scores = pd.concat([df_columns, df_scores], axis=1)
            feature_scores.columns = ['Feature_Name', 'MI Score']  # name output columns
            feature_scores.sort_values(by=['MI Score'], ascending=False, inplace=True)
            features = feature_scores.iloc[0:numFeatures]
            mutualInfoFeatures = features['Feature_Name']
            self.MIFeatures = features
            return mutualInfoFeatures
    """
    def hpTuning(self, topFeatures, batchSize, initializer, biasInit, alpha=None, gamma=None, tuner='Bayesian', MAX_TRIALS=10, EXECUTION_PER_TRIAL=2,
                 seed=42, epochs=10):
        LOG_DIR = f"{int(time.time())}"
        self.start_time = time.time()
        self.biasInit=biasInit
        self.alpha = alpha
        self.gamma = gamma
        tf.keras.backend.clear_session()

        # Set all to numpy arrays
        self.X_train = self.X_train[topFeatures].to_numpy()
        self.Y_train = self.Y_train.to_numpy()
        self.X_val = self.X_val[topFeatures].to_numpy()
        self.Y_val = self.Y_val.to_numpy()
        self.X_test = self.X_test[topFeatures].to_numpy()
        self.Y_test = self.Y_test.to_numpy()

        inputSize = self.X_train.shape[1]

        self.batch_size = batchSize


        # Class weights
        class_weights = class_weight.compute_class_weight('balanced', np.unique(self.Y_train), self.Y_train)
        class_weight_dict = dict(enumerate(class_weights))
        pos = class_weight_dict[1]
        neg = class_weight_dict[0]

        class_weight_dict[0] = 1 / self.Y_train.sum()
        class_weight_dict[1] = 1 / (len(self.Y_train) - self.Y_train.sum())

        bias = np.log(pos / neg)

        def build_model(hp):
            # define the keras model
            model = tf.keras.models.Sequential()
            model.add(tf.keras.Input(shape=(inputSize,)))

            for i in range(hp.Int('num_layers', 2, 8)):
                units = hp.Choice('units_' + str(i), values=[30, 36, 30, 41, 45, 60])
                deep_activation = hp.Choice('dense_activation_' + str(i), values=['relu', 'tanh'])

                model.add(Dense(units=units, activation=deep_activation)) #, kernel_initializer=initializer,))


                # Dropout in hidden layers
                model.add(Dropout(0.20))

                # Batch Normalization
                model.add(BatchNormalization(momentum=0.60))


            final_activation = hp.Choice('final_activation', values=['softmax', 'sigmoid'])

            if biasInit==0:
                model.add(Dense(1, activation=final_activation))
            elif biasInit==1:
                model.add(Dense(1, activation=final_activation, bias_initializer=tf.keras.initializers.Constant(value=bias)))

            # Select optimizer
            optimizer = hp.Choice('optimizer', values=['adam', 'RMSprop'])#, 'SGD'])

            lr = hp.Choice('learning_rate', [1e-3, 1e-4, 1e-5])

            # Conditional for each optimizer
            if optimizer == 'adam':
                optimizer = tf.keras.optimizers.Adam(lr, clipnorm=0.0001)
            elif optimizer == 'RMSprop':
                optimizer = tf.keras.optimizers.RMSprop(lr, clipnorm=0.0001)

            elif optimizer == 'SGD':
                optimizer = tf.keras.optimizers.SGD(lr, clipnorm=0.0001)


            if alpha!= None or gamma != None:
                loss = tfa.losses.SigmoidFocalCrossEntropy(alpha=alpha, gamma=gamma)
                self.loss = "focal_loss"
            else:
                loss = 'binary_crossentropy'
                self.loss = "binary-crossentropy"

            #fa.losses.SigmoidFocalCrossEntropy(alpha=(1/pos), gamma=0)

            # Compilation
            model.compile(optimizer=optimizer,
                          loss=weighted_loss_persample(class_weight_dict, batchSize),
                          metrics=['accuracy',
                                   tf.keras.metrics.Precision(),
                                   tf.keras.metrics.Recall(),
                                   tf.keras.metrics.AUC()])

            return model

        if tuner == 'Hyperband':
            self.hb_tuner = Hyperband(build_model,
                                          objective=kerastuner.Objective('val_auc', direction="max"),
                                          max_epochs=epochs,
                                          seed=seed,
                                          factor=3,
                                          overwrite=True)  # Set this to true if don't want to pick up from checkpoint
                                          #directory='Runs/Hyperband/Unweighted/' + LOG_DIR,
                                          #project_name='okNN' + date)

        elif tuner == 'Bayesian':
            self.hb_tuner = BayesianOptimization(build_model,
                                        objective=kerastuner.Objective('val_auc', direction="max"),
                                        overwrite=True,
                                        max_trials=MAX_TRIALS,
                                        seed=seed,
                                        executions_per_trial=2)

        elif tuner == 'Random':
            self.hb_tuner = RandomSearch(
                build_model,
                objective=kerastuner.Objective('val_auc', direction="max"),
                overwrite=True,
                seed=seed,
                max_trials=MAX_TRIALS,
                executions_per_trial=EXECUTION_PER_TRIAL,
            )

        self.hb_tuner.search(self.X_train, self.Y_train,
                                 epochs=epochs,
                                 verbose=2,
                                 validation_data=(self.X_val, self.Y_val),
                                 callbacks=[tf.keras.callbacks.EarlyStopping('val_auc', patience=4)])
                                #Early stopping will stop epochs if val_loss doesn't improve for 4 iterations

        #self.best_model = self.hb_tuner.get_best_models(num_models=1)[0]
        self.best_hps = self.hb_tuner.get_best_hyperparameters(num_trials=1)[0]
    """

    def buildModel(self, epochs):

        self.epochs = epochs

        self.best_model = self.hb_tuner.hypermodel.build(self.best_hps)
        self.history = self.best_model.fit(self.X_train, self.Y_train, epochs=epochs,
                                           validation_data=(self.X_val, self.Y_val), verbose=2)

        print(self.best_model.summary())

    def evaluateModel(self):

        wb = Workbook()
        wsResults = wb.active
        wsResults.title = "Results"
        wsHyp = wb.create_sheet(title="Hyperparameters")
        wsFeatures = wb.create_sheet(title="Features")
        wsGraphs = wb.create_sheet(title="Graphs")
        filename = self.dataset + '_' + date

        # Graphing results
        plt.clf()
        plt.cla()
        plt.close()

        plt.plot(self.history.history['auc'])
        plt.plot(self.history.history['val_auc'])
        plt.title('model auc')
        plt.ylabel('auc')
        plt.xlabel('epoch')
        plt.legend(['train', 'validation'], loc='upper right')
        figname = self.dataset + '_trainTestAUC' + date + '.png'
        plt.savefig(figname, bbox_inches="tight")
        img = openpyxl.drawing.image.Image(figname)
        img.anchor = 'A1'
        wsGraphs.add_image(img)
        #plt.show()

        plt.clf()
        plt.cla()
        plt.close()

        #plt.ylim(0.090, 0.10)
        plt.plot(self.history.history['loss'])
        plt.plot(self.history.history['val_loss'])
        plt.title('model loss')
        plt.ylabel('loss')
        plt.xlabel('epoch')
        plt.legend(['train', 'validation'], loc='upper right')
        figname = self.dataset + '_trainTestLoss' + date + '.png'
        plt.savefig(figname, bbox_inches="tight")
        img = openpyxl.drawing.image.Image(figname)
        img.anchor = 'I1'
        wsGraphs.add_image(img)
        #plt.show()

        # Feature Selection

        if self.method == 1:
            img = openpyxl.drawing.image.Image(self.dataset + 'XGBoostTopFeatures.png')
            img.anchor = 'A1'
            wsFeatures.add_image(img)

        elif self.method == 2:
            for r in dataframe_to_rows(self.Chi2features, index=False, header=True):
                wsFeatures.append(r)

        elif self.method == 3:
            for r in dataframe_to_rows(self.MIFeatures, index=False, header=True):
                wsFeatures.append(r)

        #y_predict = self.best_model.predict_classes(self.test_X)

        y_predict = (self.best_model.predict(self.X_test) > 0.5).astype("int32")

        self.specificity = specificity_score(self.Y_test, y_predict)

        self.gmean = geometric_mean_score(self.Y_test, y_predict)

        score = self.best_model.evaluate(self.X_test, self.Y_test, verbose=0)
        self.loss = score[0]
        self.accuracy = score[1]
        self.AUC = score[4]
        self.predictedNo = y_predict.sum()
        self.trueNo = self.Y_test.sum()
        self.recall = score[3]
        self.precision = score[2]
        self.tn, self.fp, self.fn, self.tp = confusion_matrix(self.Y_test, y_predict).ravel()

        results = [['loss', self.loss],
                   ['accuracy', self.accuracy],
                   ['AUC', self.AUC],
                   ['specificity', self.specificity],
                   ['recall', self.recall],
                   ['precision', self.precision],
                   ['gmean', self.gmean],
                   ['True Positive', self.tp],
                   ['True Negative', self.tn],
                   ['False Positive', self.fp],
                   ['False Negative', self.fn]]


        df = pd.DataFrame(results, columns=['Metric', 'Value'])

        for r in dataframe_to_rows(df, index=False, header=True):
            wsResults.append(r)

        #df.to_csv(filename, index=False)

        print(f'Total Cases: {len(y_predict)}')
        print(f'Predict #: {y_predict.sum()} / True # {self.Y_test.sum()}')
        print(f'True Positives #: {self.tp} / True Negatives # {self.tn}')
        print(f'False Positives #: {self.fp} / False Negatives # {self.fn}')
        print(f'Test loss: {score[0]:.6f} / Test accuracy: {score[1]:.6f} / Test AUC: {score[4]:.6f}')
        print(f'Test Recall: {score[3]:.6f} / Test Precision: {score[2]:.6f}')
        print(f'Test Specificity: {self.specificity:.6f}')
        print(f'Test Gmean: {self.gmean:.6f}')

        mins = (time.time() - self.start_time)/60  # Time in seconds

        self.best_hps.values['Batch Size'] = self.batch_size
        #self.best_hps.values['Age'] = self.age
        self.best_hps.values['epochs'] = self.epochs
        self.best_hps.values['loss'] = self.loss
        self.best_hps.values['alpha'] = self.alpha
        self.best_hps.values['gamma'] = self.gamma
        self.best_hps.values['split1'] = self.split1
        self.best_hps.values['split2'] = self.split2
        self.best_hps.values['bias'] = self.biasInit
        self.best_hps.values['time(min)'] = mins

        df = pd.DataFrame(data=self.best_hps.values, index=[0])
        df = (df.T)

        for r in dataframe_to_rows(df, index=True, header=False):
            wsHyp.append(r)

        wb.save(filename + '.xlsx')

        return self.AUC, self.gmean, self.precision, self.recall, self.specificity, self.tp, self.fp, self.tn, self.fn

class NoTune(NN):
    def buildModel(self, topFeatures, batchSize, initializer, epochs, alpha=None, gamma=None, biasInit=0):

        self.biasInit = biasInit
        self.start_time = time.time()
        self.best_hps = {'num_layers': 3,
                         'dense_activation_0': 'tanh',
                         'dense_activation_1': 'relu',
                         'dense_activation_2': 'relu',
                         'units_0': 30,
                         'units_1': 36,
                         'units_2': 45,
                         'final_activation': 'sigmoid',
                         'optimizer': 'Adam',
                         'learning_rate': 0.001}

        self.alpha = alpha
        self.gamma = gamma
        self.epochs = epochs
        LOG_DIR = f"{int(time.time())}"

        # Set all to numpy arrays
        self.X_train = self.X_train[topFeatures].to_numpy()
        self.Y_train = self.Y_train.to_numpy()
        self.X_test = self.X_test[topFeatures].to_numpy()
        self.Y_test = self.Y_test.to_numpy()

        inputSize = self.X_train.shape[1]

        self.batch_size = batchSize

        # define the keras model
        tf.keras.backend.clear_session()
        self.model = tf.keras.models.Sequential()
        self.model.add(tf.keras.Input(shape=(inputSize,)))

        # Hidden Layers
        for i in range(self.best_hps['num_layers']):
            self.model.add(
                Dense(units=self.best_hps['units_' + str(i)], activation=self.best_hps['dense_activation_' + str(i)]))
            # self.model.add(Dropout(0.20))
            # self.model.add(BatchNormalization(momentum=0.60))

        # Class weights
        class_weights = class_weight.compute_class_weight('balanced', np.unique(self.Y_train), self.Y_train)
        class_weight_dict = dict(enumerate(class_weights))
        pos = class_weight_dict[1]
        neg = class_weight_dict[0]

        bias = np.log(pos / neg)

        if biasInit == 0:
            # Final Layer
            self.model.add(Dense(1, activation=self.best_hps['final_activation']))
        elif biasInit == 1:
            # Final Layer
            self.model.add(Dense(1, activation=self.best_hps['final_activation'],
                                 bias_initializer=tf.keras.initializers.Constant(
                                     value=bias)))  # kernel_initializer=initializer,

        # Loss Function
        if alpha != None or gamma != None:
            loss = tfa.losses.SigmoidFocalCrossEntropy(alpha=alpha, gamma=gamma)
            self.loss = "focal_loss"
        else:
            loss = 'binary_crossentropy'
            self.loss = "binary-crossentropy"

        # Compilation
        self.model.compile(optimizer=tf.keras.optimizers.Adam(learning_rate=self.best_hps['learning_rate']),
                           loss=loss,
                           metrics=['accuracy',
                                    tf.keras.metrics.Precision(),
                                    tf.keras.metrics.Recall(),
                                    tf.keras.metrics.AUC()])

        self.history = self.model.fit(self.X_train, self.Y_train, epochs=epochs, validation_data=(self.X_val, self.Y_val),
                                      verbose=2, class_weight=class_weight_dict)
    def evaluateModel(self):

        wb = Workbook()
        wsResults = wb.active
        wsResults.title = "Results"
        wsHyp = wb.create_sheet(title="Hyperparameters")
        wsFeatures = wb.create_sheet(title="Features")
        wsGraphs = wb.create_sheet(title="Graphs")
        filename = self.dataset + '_' + date

        # Graphing results
        plt.clf()
        plt.cla()
        plt.close()

        # plt.ylim(0.40, 0.66)
        plt.plot(self.history.history['auc'])
        plt.plot(self.history.history['val_auc'])
        plt.title('model auc')
        plt.ylabel('auc')
        plt.xlabel('epoch')
        plt.legend(['train', 'validation'], loc='upper right')
        figname = self.dataset + '_trainTestAUC' + date + '.png'
        plt.savefig(figname, bbox_inches="tight")
        img = openpyxl.drawing.image.Image(figname)
        img.anchor = 'A1'
        #wsGraphs.add_image(img)
        # plt.show()

        plt.clf()
        plt.cla()
        plt.close()

        # plt.ylim(0.0, 0.15)
        plt.plot(self.history.history['loss'])
        plt.plot(self.history.history['val_loss'])
        plt.title('model loss')
        plt.ylabel('loss')
        plt.xlabel('epoch')
        plt.legend(['train', 'validation'], loc='upper right')
        figname = self.dataset + '_trainTestLoss' + date + '.png'
        plt.savefig(figname, bbox_inches="tight")
        img = openpyxl.drawing.image.Image(figname)
        img.anchor = 'J1'
        #wsGraphs.add_image(img)
        # plt.show()

        # Feature Selection
        if self.method == 1:
            img = openpyxl.drawing.image.Image(self.dataset + 'XGBoostTopFeatures.png')
            img.anchor = 'A1'
            wsFeatures.add_image(img)

        elif self.method == 2:
            for r in dataframe_to_rows(self.Chi2features, index=False, header=True):
                wsFeatures.append(r)

        elif self.method == 3:
            for r in dataframe_to_rows(self.MIFeatures, index=False, header=True):
                wsFeatures.append(r)


        y_predict = (self.model.predict(self.X_test) > 0.5).astype("int32")

        self.specificity = specificity_score(self.Y_test, y_predict)

        self.gmean = geometric_mean_score(self.Y_test, y_predict)

        score = self.model.evaluate(self.X_test, self.Y_test, verbose=0)
        self.loss = score[0]
        self.accuracy = score[1]
        self.AUC = score[4]
        self.predictedNo = y_predict.sum()
        self.trueNo = self.Y_test.sum()
        self.recall = score[3]
        self.precision = score[2]
        self.tn, self.fp, self.fn, self.tp = confusion_matrix(self.Y_test, y_predict).ravel()

        results = [['loss', self.loss],
                   ['accuracy', self.accuracy],
                   ['AUC', self.AUC],
                   ['specificity', self.specificity],
                   ['recall', self.recall],
                   ['precision', self.precision],
                   ['gmean', self.gmean],
                   ['True Positive', self.tp],
                   ['True Negative', self.tn],
                   ['False Positive', self.fp],
                   ['False Negative', self.fn]]

        df = pd.DataFrame(results, columns=['Metric', 'Value'])


        for r in dataframe_to_rows(df, index=False, header=True):
            wsResults.append(r)

        # df.to_csv(filename, index=False)

        print(f'Total Cases: {len(y_predict)}')
        print(f'Predict #: {y_predict.sum()} / True # {self.Y_test.sum()}')
        print(f'True Positives #: {self.tp} / True Negatives # {self.tn}')
        print(f'False Positives #: {self.fp} / False Negatives # {self.fn}')
        print(f'Test loss: {score[0]:.6f} / Test accuracy: {score[1]:.6f} / Test AUC: {score[4]:.6f}')
        print(f'Test Recall: {score[3]:.6f} / Test Precision: {score[2]:.6f}')
        print(f'Test Specificity: {self.specificity:.6f}')
        print(f'Test Gmean: {self.gmean:.6f}')

        mins = (time.time() - self.start_time) / 60  # Time in seconds

        self.best_hps['Batch Size'] = self.batch_size
        self.best_hps['Age'] = self.age
        self.best_hps['epochs'] = self.epochs
        self.best_hps['loss'] = self.loss
        self.best_hps['alpha'] = self.alpha
        self.best_hps['gamma'] = self.gamma
        self.best_hps['split1'] = self.split1
        self.best_hps['split2'] = self.split2
        self.best_hps['bias'] = self.biasInit
        self.best_hps['time(mins)'] = mins

        df = pd.DataFrame(data=self.best_hps, index=[0])
        df = (df.T)
        for r in dataframe_to_rows(df, index=True, header=False):
            wsHyp.append(r)

        wb.save(filename + '.xlsx')

        return self.AUC, self.gmean, self.precision, self.recall, self.specificity, self.tp, self.fp, self.tn, self.fn, self.loss

if __name__ == "__main__":



    # Get data
    parent = os.path.dirname(os.getcwd())
    dataPath = os.path.join(parent, 'Data/Processed/Oklahoma/Complete/Full/Outliers/Chi2_Categorical.csv')


    resultPath = os.path.join(parent, 'Results/Oklahoma/Full/NoTune/ClassWeights/Test_a5g175')


    modelWeighted = NoTune(resultPath)
    modelWeighted.prepData(age='Categorical',
                           data=dataPath)
    modelWeighted.splitData(testSize=0.20, valSize=0.20)
    # modelWeighted.detectOutliers('lof', con=0.1)
    features = modelWeighted.featureSelection(numFeatures=20, method=2)

    """
    modelWeighted.hpTuning(features, batchSize=4096, tuner='Hyperband',
                           initializer='RandomUniform', biasInit=1,
                           MAX_TRIALS=5, epochs=5)
    # For hypertuning
    modelWeighted.buildModel(epochs=30)
    """

    # For hand-tuning
    modelWeighted.buildModel(features, batchSize=4096, alpha=0.5, gamma=1.75, initializer='RandomUniform', biasInit=1, epochs=30)
    auc, gmean, precision, recall, specificity, tp, fp, tn, fn, loss = modelWeighted.evaluateModel()


