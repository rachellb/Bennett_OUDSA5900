#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# For handling data
import pandas as pd
import numpy as np
from datetime import datetime

from sklearn.preprocessing import OrdinalEncoder
from sklearn.utils import class_weight
from statistics import mean

# For imputing data
from sklearn.experimental import enable_iterative_imputer
from sklearn.impute import IterativeImputer

# For feature selection
from sklearn.feature_selection import SelectKBest, chi2
from xgboost import XGBClassifier
from matplotlib import pyplot
from xgboost import plot_importance
from sklearn.metrics import accuracy_score
from sklearn.model_selection import train_test_split
from sklearn.feature_selection import mutual_info_classif
from sklearn import preprocessing
from sklearn.feature_selection import f_classif
# import hyperopt
# from hyperopt import fmin, tpe, hp, STATUS_OK, Trials

# For balancing batches
from imblearn.keras import BalancedBatchGenerator
from imblearn.over_sampling import RandomOverSampler

# For NN and tuning
import tensorflow as tf
import kerastuner
from kerastuner.tuners import Hyperband, BayesianOptimization
from tensorflow.keras.layers import Dense, Dropout, BatchNormalization

# For additional metrics
from imblearn.metrics import geometric_mean_score, specificity_score
from sklearn.metrics import confusion_matrix
import tensorflow_addons as tfa  # For focal loss function
import time
import matplotlib.pyplot as plt
from openpyxl import Workbook  # For storing results
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
from openpyxl import load_workbook
import os
date = datetime.today().strftime('%m%d%y')  # For labelling purposes
from NeuralNetworkBase import NN
from Cleaning.Clean import *


def weighted_loss_persample(weights, batchSize):
    def loss(y_true, y_pred):
        # The masks for the true and false values

        idx_1 = y_true == 1.
        idx_0 = y_true == 0.

        pred_1 = tf.boolean_mask(y_pred, idx_1)
        pred_1 = tf.expand_dims(pred_1, 1)

        true_1 = tf.boolean_mask(y_true, idx_1)
        true_1 = tf.expand_dims(true_1, 1)

        pred_0 = tf.boolean_mask(y_pred, idx_0)
        pred_0 = tf.expand_dims(pred_0, 1)

        true_0 = tf.boolean_mask(y_true, idx_0)
        true_0 = tf.expand_dims(true_0, 1)

        bce = tf.keras.losses.BinaryCrossentropy(reduction=tf.keras.losses.Reduction.NONE)

        # The losses for the 0 classes
        losses_0 = bce(pred_0, true_0)
        losses_0 = tf.reduce_sum(losses_0, 0)# Add back up
        losses_0 = losses_0 * weights[0]


        # The losses for the 1 classes
        losses_1 = bce(pred_1, true_1)
        losses_1 = tf.reduce_sum(losses_1, 0)  # Add back up
        losses_1 = losses_1 * weights[1]

        # Add them back up and divide by batch size
        sum = losses_0 + losses_1
        total = sum/batchSize

        # anything you like to print
        print_op = tf.print("\nDebug : ", tf.shape(y_pred))
        with tf.control_dependencies([print_op]):  # this will force the print_op to be run
            return total


    return loss


class fullNN(NN):
    def __init__(self, PARAMS, dataset):

        self.PARAMS = PARAMS
        self.dataset = dataset


    def prepData(self, age, data):

        self.age=age

        self.data = pd.read_csv(data)

class NoTune(fullNN):


    def buildModel(self, topFeatures, batchSize, initializer, epochs, alpha=None, gamma=None, biasInit=0):

        self.biasInit=biasInit
        self.start_time = time.time()
        self.best_hps = {'num_layers': 3,
                         'dense_activation_0': 'tanh',
                         'dense_activation_1': 'relu',
                         'dense_activation_2': 'relu',
                         'units_0': 30,
                         'units_1': 36,
                         'units_2': 45,
                         'final_activation': 'sigmoid',
                         'optimizer': 'Adam',
                         'learning_rate': 0.001}

        self.alpha = alpha
        self.gamma = gamma
        self.epochs = epochs
        LOG_DIR = f"{int(time.time())}"

        # Set all to numpy arrays
        self.X_train = self.X_train[topFeatures].to_numpy()
        self.Y_train = self.Y_train.to_numpy()
        self.X_val = self.X_val[topFeatures].to_numpy()
        self.Y_val = self.Y_val.to_numpy()
        self.X_test = self.X_test[topFeatures].to_numpy()
        self.Y_test = self.Y_test.to_numpy()

        inputSize = self.X_train.shape[1]

        self.batch_size = batchSize

        self.training_generator = BalancedBatchGenerator(self.X_train, self.Y_train,
                                                         batch_size=self.batch_size,
                                                         sampler=RandomOverSampler(),
                                                         random_state=42)
        self.validation_generator = BalancedBatchGenerator(self.X_val, self.Y_val,
                                                         batch_size=self.batch_size,
                                                         sampler=RandomOverSampler(),
                                                         random_state=42)

        # define the keras model
        tf.keras.backend.clear_session()
        self.model = tf.keras.models.Sequential()
        self.model.add(tf.keras.Input(shape=(inputSize,)))


        # Hidden Layers
        for i in range(self.best_hps['num_layers']):
            self.model.add(Dense(units=self.best_hps['units_' + str(i)], activation=self.best_hps['dense_activation_' + str(i)]))
            self.model.add(Dropout(0.20))
            #self.model.add(BatchNormalization(momentum=0.60))



        # Class weights
        class_weights = class_weight.compute_class_weight('balanced', np.unique(self.Y_train), self.Y_train)
        class_weight_dict = dict(enumerate(class_weights))
        pos = class_weight_dict[1]
        neg = class_weight_dict[0]

        bias = np.log(pos/neg)


        if biasInit==0:
            # Final Layer
            self.model.add(Dense(1, activation=self.best_hps['final_activation']))
        elif biasInit==1:
            # Final Layer
            self.model.add(Dense(1, activation=self.best_hps['final_activation'], bias_initializer=tf.keras.initializers.Constant(value=bias)))#kernel_initializer=initializer,

        # Loss Function
        if alpha != None or gamma != None:
            loss = tfa.losses.SigmoidFocalCrossEntropy(alpha=alpha, gamma=gamma)
            self.loss = "focal_loss"
        else:
            loss = 'binary_crossentropy'
            self.loss = "binary-crossentropy"

        # Compilation
        self.model.compile(optimizer=tf.keras.optimizers.Adam(learning_rate=self.best_hps['learning_rate']),
                      loss=loss,
                      metrics=['accuracy',
                               tf.keras.metrics.Precision(),
                               tf.keras.metrics.Recall(),
                               tf.keras.metrics.AUC()])


        self.history = self.model.fit(self.training_generator, epochs=epochs, validation_data=self.validation_generator,
                                      verbose=2, class_weight=class_weight_dict)

    def evaluateModel(self):
        wb = Workbook()
        wsResults = wb.active
        wsResults.title = "Results"
        wsHyp = wb.create_sheet(title="Hyperparameters")
        wsFeatures = wb.create_sheet(title="Features")
        wsGraphs = wb.create_sheet(title="Graphs")
        filename = self.dataset + '_' + date

        # Graphing results
        plt.clf()
        plt.cla()
        plt.close()

        #plt.ylim(0.40, 0.66)
        plt.plot(self.history.history['auc'])
        plt.plot(self.history.history['val_auc'])
        plt.title('model auc')
        plt.ylabel('auc')
        plt.xlabel('epoch')
        plt.legend(['train', 'validation'], loc='upper right')
        figname=self.dataset + '_trainTestAUC' + date + '.png'
        plt.savefig(figname, bbox_inches="tight")
        img = openpyxl.drawing.image.Image(figname)
        img.anchor = 'A1'
        wsGraphs.add_image(img)
        # plt.show()

        plt.clf()
        plt.cla()
        plt.close()

        #plt.ylim(0.0, 0.15)
        plt.plot(self.history.history['loss'])
        plt.plot(self.history.history['val_loss'])
        plt.title('model loss')
        plt.ylabel('loss')
        plt.xlabel('epoch')
        plt.legend(['train', 'validation'], loc='upper right')
        figname=self.dataset + '_trainTestLoss' + date +'.png'
        plt.savefig(figname, bbox_inches="tight")
        img = openpyxl.drawing.image.Image(figname)
        img.anchor = 'J1'
        wsGraphs.add_image(img)
        # plt.show()



        # Feature Selection
        if self.method == 1:
            img = openpyxl.drawing.image.Image(self.dataset + 'XGBoostTopFeatures.png')
            img.anchor = 'A1'
            wsFeatures.add_image(img)

        elif self.method == 2:
            for r in dataframe_to_rows(self.Chi2features, index=False, header=True):
                wsFeatures.append(r)

        elif self.method == 3:
            for r in dataframe_to_rows(self.MIFeatures, index=False, header=True):
                wsFeatures.append(r)


        # y_predict = self.best_model.predict_classes(self.test_X) # deprecated

        y_predict = (self.model.predict(self.X_test) > 0.5).astype("int32")

        self.specificity = specificity_score(self.Y_test, y_predict)

        self.gmean = geometric_mean_score(self.Y_test, y_predict)

        score = self.model.evaluate(self.X_test, self.Y_test, verbose=0)
        self.loss = score[0]
        self.accuracy = score[1]
        self.AUC = score[4]
        self.predictedNo = y_predict.sum()
        self.trueNo = self.Y_test.sum()
        self.recall = score[3]
        self.precision = score[2]
        self.tn, self.fp, self.fn, self.tp = confusion_matrix(self.Y_test, y_predict).ravel()

        results = [['loss', self.loss],
                   ['accuracy', self.accuracy],
                   ['AUC', self.AUC],
                   ['specificity', self.specificity],
                   ['recall', self.recall],
                   ['precision', self.precision],
                   ['gmean', self.gmean],
                   ['True Positive', self.tp],
                   ['True Negative', self.tn],
                   ['False Positive', self.fp],
                   ['False Negative', self.fn]]

        df = pd.DataFrame(results, columns=['Metric', 'Value'])

        for r in dataframe_to_rows(df, index=False, header=True):
            wsResults.append(r)

        # df.to_csv(filename, index=False)

        print(f'Total Cases: {len(y_predict)}')
        print(f'Predict #: {y_predict.sum()} / True # {self.Y_test.sum()}')
        print(f'True Positives #: {self.tp} / True Negatives # {self.tn}')
        print(f'False Positives #: {self.fp} / False Negatives # {self.fn}')
        print(f'Test loss: {score[0]:.6f} / Test accuracy: {score[1]:.6f} / Test AUC: {score[4]:.6f}')
        print(f'Test Recall: {score[3]:.6f} / Test Precision: {score[2]:.6f}')
        print(f'Test Specificity: {self.specificity:.6f}')
        print(f'Test Gmean: {self.gmean:.6f}')

        mins = (time.time() - self.start_time) / 60  # Time in seconds

        self.best_hps['Batch Size'] = self.batch_size
        self.best_hps['Age'] = self.age
        self.best_hps['epochs'] = self.epochs
        self.best_hps['loss'] = self.loss
        self.best_hps['alpha'] = self.alpha
        self.best_hps['gamma'] = self.gamma
        self.best_hps['split1'] = self.split1
        self.best_hps['split2'] = self.split2
        self.best_hps['bias'] = self.biasInit
        self.best_hps['time(mins)'] = mins

        df = pd.DataFrame(data=self.best_hps, index=[0])
        df = (df.T)
        for r in dataframe_to_rows(df, index=True, header=False):
            wsHyp.append(r)

        wb.save(filename + '.xlsx')

        return self.AUC, self.gmean, self.precision, self.recall, self.specificity, self.tp, self.fp, self.tn, self.fn, self.loss

class NoGen(NoTune):

    def buildModel(self, topFeatures, batchSize, initializer, epochs, alpha=None, gamma=None, biasInit=0):
        self.biasInit = biasInit
        self.start_time = time.time()
        self.best_hps = {'num_layers': 3,
                         'dense_activation_0': 'tanh',
                         'dense_activation_1': 'relu',
                         'dense_activation_2': 'relu',
                         'units_0': 30,
                         'units_1': 36,
                         'units_2': 45,
                         'final_activation': 'sigmoid',
                         'optimizer': 'Adam',
                         'learning_rate': 0.001}

        self.alpha = alpha
        self.gamma = gamma
        self.epochs = epochs
        LOG_DIR = f"{int(time.time())}"

        # Set all to numpy arrays
        self.X_train = self.X_train[topFeatures]
        self.Y_train = self.Y_train
        self.X_val = self.X_val[topFeatures]
        self.Y_val = self.Y_val
        self.X_test = self.X_test[topFeatures]
        self.Y_test = self.Y_test

        inputSize = self.X_train.shape[1]

        self.batch_size = batchSize

        # define the keras model
        tf.keras.backend.clear_session()
        self.model = tf.keras.models.Sequential()
        self.model.add(tf.keras.Input(shape=(inputSize,)))

        # Hidden Layers
        for i in range(self.best_hps['num_layers']):
            self.model.add(
                Dense(units=self.best_hps['units_' + str(i)], activation=self.best_hps['dense_activation_' + str(i)]))
            self.model.add(Dropout(0.20))
            self.model.add(BatchNormalization(momentum=0.60))

        # Class weights
        class_weights = class_weight.compute_class_weight('balanced', np.unique(self.Y_train), self.Y_train)
        class_weight_dict = dict(enumerate(class_weights))
        pos = class_weight_dict[1]
        neg = class_weight_dict[0]

        scalar = len(self.Y_train)
        class_weight_dict[0] = scalar / self.Y_train.value_counts()[0]
        class_weight_dict[1] = scalar / self.Y_train.value_counts()[1]



        bias = np.log(pos / neg)

        if biasInit == 0:
            # Final Layer
            self.model.add(Dense(1, activation=self.best_hps['final_activation']))
        elif biasInit == 1:
            # Final Layer
            self.model.add(Dense(1, activation=self.best_hps['final_activation'],
                                 bias_initializer=tf.keras.initializers.Constant(
                                     value=bias)))  # kernel_initializer=initializer,

        # Loss Function
        if alpha != None or gamma != None:
            loss = tfa.losses.SigmoidFocalCrossEntropy(alpha=alpha, gamma=gamma)
            self.loss = "focal_loss"
        else:
            loss = 'binary_crossentropy'
            self.loss = "binary-crossentropy"

        # Compilation
        self.model.compile(optimizer=tf.keras.optimizers.Adam(learning_rate=self.best_hps['learning_rate']),
                           loss=weighted_loss_persample(class_weight_dict, batchSize),
                           metrics=['accuracy',
                                    tf.keras.metrics.Precision(),
                                    tf.keras.metrics.Recall(),
                                    tf.keras.metrics.AUC()])

        self.history = self.model.fit(self.X_train, self.Y_train, batch_size=batchSize,
                                      epochs=epochs, validation_data=(self.X_val, self.Y_val),
                                      verbose=2, class_weight=class_weight_dict)

if __name__ == "__main__":

    PARAMS = {'lr': 0.005,
              'momentum': 0.9,
              'epochs': 10,
              'batch_size': 4096}

    # Get data
    parent = os.path.dirname(os.getcwd())
    path = os.path.join(parent, 'Data/Processed/Oklahoma/Complete/Full/Outliers/Chi2_Categorical.csv')


    filename = 'Results/Oklahoma/Full/NoTune/ClassWeights/Test_a5g175'
    modelWeighted = fullNN(PARAMS, filename)
    modelWeighted.prepData(age='Categorical',
                           data=path)
    modelWeighted.splitData(testSize=0.10, valSize=0.10)
    # modelWeighted.detectOutliers('lof', con=0.1)
    features = modelWeighted.featureSelection(numFeatures=20, method=2)

    """
    modelWeighted.hpTuning(features, batchSize=2048, tuner='Hyperband',
                           initializer='RandomUniform', biasInit=0,
                           MAX_TRIALS=5, epochs=5)
    """

    # For hypertuning
    # modelWeighted.buildModel(epochs=30)

    # For hand-tuning
    modelWeighted.buildModel(features, batchSize=PARAMS['batch_size'], alpha=0.5, gamma=1.75, initializer='RandomUniform',
                             biasInit=1, epochs=30)

    auc, gmean, precision, recall, specificity, tp, fp, tn, fn, loss = modelWeighted.evaluateModel()



    def storeResults():
        wb = Workbook()
        dest_filename = 'Results/Texas/Full/NoTune/ClassWeights/FocalTests/Results_AlphaGamma_Full.xlsx'


        #initializers = ['RandomUniform', 'RandomNormal', 'he_uniform', 'he_normal', 'glorot_uniform', 'glorot_normal']

        ws = wb.active
        ws['B1'] = "AUC"
        ws['C1'] = "Gmean"
        ws['D1'] = "Precision"
        ws['E1'] = "Recall"
        ws['F1'] = "Specificity"

        # This section is for testing out different combinations of alpha and gamma

        wsAUC = wb.active
        wsAUC.title = "AUC"
        wsGmean = wb.create_sheet(title="Gmean")
        wsPrecision = wb.create_sheet(title="Precision")
        wsRecall = wb.create_sheet(title="Recall")
        wsSpecificity = wb.create_sheet(title="Specificity")
        wstp = wb.create_sheet(title="True Positive")
        wstn = wb.create_sheet(title="True Negative")
        wsfp = wb.create_sheet(title="False Positive")
        wsfn = wb.create_sheet(title="False Negative")


        #Filling index and columns
        ch = 'B'
        alpha = 0.25
        gamma = 1
        for j in range(5):
            wsAUC[chr(ord(ch) + j) + str(1)] = "Gamma= " + str(gamma)
            wsGmean[chr(ord(ch) + j) + str(1)] = "Gamma= " + str(gamma)
            wsPrecision[chr(ord(ch) + j) + str(1)] = "Gamma= " + str(gamma)
            wsRecall[chr(ord(ch) + j) + str(1)] = "Gamma= " + str(gamma)
            wsSpecificity[chr(ord(ch) + j) + str(1)] = "Gamma= " + str(gamma)
            wstp[chr(ord(ch) + j) + str(1)] = "Gamma= " + str(gamma)
            wstn[chr(ord(ch) + j) + str(1)] = "Gamma= " + str(gamma)
            wsfp[chr(ord(ch) + j) + str(1)] = "Gamma= " + str(gamma)
            wsfn[chr(ord(ch) + j) + str(1)] = "Gamma= " + str(gamma)

            gamma = gamma + 0.25

        for i in range(4):
            wsAUC['A' + str(i+2)] = "Alpha= " + str(alpha)
            wsGmean['A' + str(i+2)] = "Alpha= " + str(alpha)
            wsPrecision['A' + str(i+2)] = "Alpha= " + str(alpha)
            wsRecall['A' + str(i+2)] = "Alpha= " + str(alpha)
            wsSpecificity['A' + str(i+2)] = "Alpha= " + str(alpha)
            wstp['A' + str(i + 2)] = "Alpha= " + str(alpha)
            wstn['A' + str(i + 2)] = "Alpha= " + str(alpha)
            wsfp['A' + str(i + 2)] = "Alpha= " + str(alpha)
            wsfn['A' + str(i + 2)] = "Alpha= " + str(alpha)

            alpha = alpha + 0.25


        ch = 'B' #Start at column B
        alpha = 0.25
        gamma = 1


        for i in range(4):
            # Reset gamma at the end of each loop
            gamma = 1
            for j in range(5):
                filename = 'Results/Texas/Full/NoTune/ClassWeights/NoGen_30_a' + str(alpha) + "g" + str(gamma)
                modelWeighted = NoGen(filename)
                modelWeighted.prepData(age='Categorical',
                                       data='Data/Processed/Texas/Full/Outliers/Complete/Chi2_Categorical.csv')
                modelWeighted.splitData(testSize=0.10, valSize=0.10)
                # modelWeighted.detectOutliers('lof', con=0.1)
                features = modelWeighted.featureSelection(numFeatures=20, method=2)


                modelWeighted.hpTuning(features, batchSize=8192, tuner='Hyperband',
                                       initializer='RandomUniform', biasInit=1,
                                       MAX_TRIALS=5, epochs=30)


                # For hypertuning
                # modelWeighted.buildModel(epochs=30)

                # For hand-tuning
                modelWeighted.buildModel(features, batchSize=8192, alpha=alpha, gamma=gamma, initializer='RandomUniform',
                                         biasInit=1, epochs=30)
                auc, gmean, precision, recall, specificity, tp, fp, tn, fn, loss = modelWeighted.evaluateModel()

                wsAUC[chr(ord(ch) + j) + str(i+2)] = auc
                wsGmean[chr(ord(ch) + j) + str(i+2)] = gmean
                wsPrecision[chr(ord(ch) + j) + str(i+2)] = precision
                wsRecall[chr(ord(ch) + j) + str(i+2)] = recall
                wsSpecificity[chr(ord(ch) + j) + str(i+2)] = specificity
                wstp[chr(ord(ch) + j) + str(i + 2)] = tp
                wsfp[chr(ord(ch) + j) + str(i + 2)] = fp
                wstn[chr(ord(ch) + j) + str(i + 2)] = tn
                wsfn[chr(ord(ch) + j) + str(i + 2)] = fn

                gamma = gamma + 0.25
            alpha = alpha + 0.25



        wb.save(filename=dest_filename)

